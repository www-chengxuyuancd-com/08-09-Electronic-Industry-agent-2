#!/usr/bin/env python3
"""
Python Backend for Electronic Industry Agent
FastAPI-based backend replacing Next.js API routes
"""

import os
import json
import re
import asyncio
from datetime import datetime
from typing import Dict, Any, List, Optional, AsyncGenerator, Tuple, Set
from contextlib import asynccontextmanager
import uuid
import csv
import aiofiles
import pathlib
import hashlib
import traceback

import asyncpg
import pandas as pd
import httpx
from fastapi import FastAPI, HTTPException, BackgroundTasks, UploadFile, File
from fastapi.middleware.cors import CORSMiddleware
from fastapi.responses import StreamingResponse, FileResponse
from pydantic import BaseModel, Field
from dotenv import load_dotenv
from openpyxl.styles import PatternFill
from openpyxl.utils import get_column_letter

# Load environment variables
load_dotenv()
# Config module import (dataset presets)
try:
    import sys as _sys
    _cfg_dir = os.path.join(os.path.dirname(__file__), "electronic-industry-agent")
    if _cfg_dir not in _sys.path:
        _sys.path.append(_cfg_dir)
    from config import DATASET_PRESETS as CONFIG_DATASET_PRESETS
except Exception:
    CONFIG_DATASET_PRESETS = {}

# Database configuration
DATABASE_URL = os.getenv("DATABASE_URL", "postgresql://username:password@localhost:5432/dbname")

# LLM configuration
OPENAI_API_KEY = os.getenv("OPENAI_API_KEY")
OPENAI_API_ENDPOINT = os.getenv("OPENAI_API_ENDPOINT", "https://api.openai.com/v1/chat/completions")
OPENAI_MODEL = os.getenv("OPENAI_MODEL", "gpt-3.5-turbo")

GEMINI_API_KEY = os.getenv("GEMINI_API_KEY")
GEMINI_ENDPOINT = os.getenv("GEMINI_ENDPOINT", "https://generativelanguage.googleapis.com/v1beta/models/gemini-2.0-flash-exp:generateContent")

# DeepSeek (OpenAI-compatible Chat Completions)
DEEPSEEK_API_KEY = os.getenv("DEEPSEEK_API_KEY")
DEEPSEEK_API_ENDPOINT = os.getenv("DEEPSEEK_API_ENDPOINT", "https://api.deepseek.com/v1/chat/completions")
DEEPSEEK_MODEL = os.getenv("DEEPSEEK_MODEL", "deepseek-chat")

LLM_PROVIDER = os.getenv("LLM_PROVIDER", "deepseek")

# LLM HTTP settings
LLM_HTTP_TIMEOUT_SECONDS = int(os.getenv("LLM_HTTP_TIMEOUT_SECONDS", "30"))
LLM_MAX_RETRIES = int(os.getenv("LLM_MAX_RETRIES", "2"))
LLM_BACKOFF_BASE = float(os.getenv("LLM_BACKOFF_BASE", "0.8"))

# Database schema (same as from the original TypeScript file)
DB_SCHEMA = """
```sql
create schema assessment;

create schema assessment_profile;

create schema platform_management;

create table platform_management.assignment
(
    id                    bigserial
        primary key,
    assignment_id         integer
        constraint uk_69g7kc6aba38ao4e3el6wmf6l
            unique,
    effort                integer,
    assignment_end_date   date,
    opportunity_id        varchar(255),
    shadow                boolean,
    staffing_request_id   varchar(255),
    assignment_start_date date,
    status                boolean,
    employee_id           varchar(255)
);

create table platform_management."user"
(
    id          bigserial
        primary key,
    created_at  timestamp,
    email       varchar(255)
        constraint uk_ob8kqyqqgmefl0aco34akdtpe
            unique,
    employee_id varchar(255)
        constraint uk_r1usl9qoplqsbrhha5e0niqng
            unique,
    is_active   boolean,
    name        varchar(255),
    updated_at  timestamp
);

-- Additional tables truncated for brevity, full schema available in doc
```
"""

# SQL prompt (truncated for brevity)
SQL_PROMPT = """
企业级BI数据示例，涵盖常见的复杂数据图表场景。
请生成PostgreSQL兼容的SQL语句，不包含任何解释或代码块标记。
"""

# Prompt helpers
def build_sql_prompt(user_input: str) -> str:
    instructions = (
        "你是资管与网管数据分析专家。根据用户需求，输出严格的PostgreSQL查询语句。"
        "注意：不要输出除SQL以外的任何文字，不要包裹Markdown代码块。"
    )
    return f"{instructions}\n业务背景:\n{SQL_PROMPT}\n用户需求:\n{user_input}\n只输出SQL："

def build_thinking_prompt(user_input: str) -> str:
    instructions = (
        "你是资管与网管数据分析专家。请先简要给出思考过程，再给出SQL语句。"
        "输出格式必须包含两段：\n思考过程：<你的简要思考>\nSQL语句：<仅SQL>"
    )
    return f"{instructions}\n业务背景:\n{SQL_PROMPT}\n用户需求:\n{user_input}"

def build_intent_prompt(user_input: str) -> str:
    instructions = """
你是一个意图识别助手。根据用户输入，从以下列表中识别需要执行的任务，并抽取必要的参数。
- 任务类型: 'OLT_STATISTICS' (OLT统计) 或 'FTTR_CHECK' (FTTR鉴别)。
- 当识别为 FTTR_CHECK 时，尽量抽取以下参数(若有)：
  - erjiFenGuang: 二级分光器名称（原文字符串，允许包含中文括号与'/'）
  - onuMingCheng: ONU用户名称（原文字符串）
输出严格的 JSON，禁止任何解释或多余文本，结构如下：
{
  "tasks": [
    {"type": "FTTR_CHECK", "params": {"erjiFenGuang": "...", "onuMingCheng": "..."}},
    {"type": "OLT_STATISTICS", "params": {}}
  ]
}
注意：
- 保持 tasks 为数组，可包含 0-N 个任务。
- 字段名必须使用上述英文字段。
- 参数缺失时用空对象或省略该字段。
- 若文本显式提到 ONU 名称（常和"ONU用户"或 ONU 一起出现且带引号），优先填写 onuMingCheng。
"""
    return f"{instructions}\n用户输入:\n{user_input}\n只输出JSON："

# Global database connection pool
db_pool: Optional[asyncpg.Pool] = None

# 应用生命周期管理器：负责启动时创建数据库连接池，关闭时安全释放，保证后台任务可用性
@asynccontextmanager
async def lifespan(app: FastAPI):
    """Application lifespan manager"""
    global db_pool
    
    # Initialize database connection pool
    try:
        db_pool = await asyncpg.create_pool(DATABASE_URL)
        print("Database connection pool created successfully")
    except Exception as e:
        print(f"Failed to create database pool: {e}")
        db_pool = None
    
    yield
    
    # Cleanup
    if db_pool:
        await db_pool.close()
        print("Database connection pool closed")

# FastAPI app
# 创建 FastAPI 应用：配置标题、描述、版本与生命周期钩子
app = FastAPI(
    title="Electronic Industry Agent Backend",
    description="Python backend for Electronic Industry Agent",
    version="1.0.0",
    lifespan=lifespan
)

# CORS middleware
dev_allowed_origins = [
    "http://localhost:3000",
    "http://127.0.0.1:3000",
    "http://localhost:3001",
    "http://127.0.0.1:3001",
]

# 跨域设置：允许本地与局域网调试访问，方法与头部均放开
app.add_middleware(
    CORSMiddleware,
    allow_origins=dev_allowed_origins,
    # Allow common LAN hosts like 192.168.x.x:port during development
    allow_origin_regex=r"http://(localhost|127\\.0\\.0\\.1|192\\.168\\.[0-9]{1,3}\\.[0-9]{1,3}|10\\.[0-9]{1,3}\\.[0-9]{1,3}\\.[0-9]{1,3}|172\\.(1[6-9]|2[0-9]|3[0-1])\\.[0-9]{1,3}\\.[0-9]{1,3}):[0-9]+",
    allow_credentials=True,
    allow_methods=["*"],
    allow_headers=["*"],
)

# Pydantic models
# 模型定义：用于接口的请求/响应校验与自动文档生成，便于前后端契约对齐
class LLMRequest(BaseModel):
    userInput: str = Field(..., description="User input for LLM processing")
    modelType: str = Field(default="deepseek", description="LLM model type")

class SQLQueryRequest(BaseModel):
    sql: str = Field(..., description="SQL query to execute")
    # Optional entity hints to enrich empty-state UX
    entityType: Optional[str] = Field(default=None, description="Entity type, e.g., 'ONU' or '分光器'")
    entityName: Optional[str] = Field(default=None, description="Entity name extracted from user input")

class LLMStreamRequest(BaseModel):
    userInput: str = Field(..., description="User input for streaming LLM")

class IntentLLMRequest(BaseModel):
    text: str
    provider: Optional[str] = None

class ErrorResponse(BaseModel):
    success: bool = False
    error: Dict[str, str]

class SuccessResponse(BaseModel):
    sql: str

class IntentRequest(BaseModel):
    text: str = Field(..., description="用户输入的原始文本")

class IntentTask(BaseModel):
    type: str = Field(..., description="任务类型，如 OLT_STATISTICS 或 FTTR_CHECK")
    params: Dict[str, Any] = Field(default_factory=dict, description="任务需要的参数")

class IntentResponse(BaseModel):
    tasks: List[IntentTask]

class OLTStatisticsResponse(BaseModel):
    preview: List[Dict[str, Any]]
    fileId: str
    filename: str
    downloadUrl: str
    rowCount: int

class FTTRCheckRequest(BaseModel):
    erjiFenGuang: Optional[str] = Field(default=None, description="二级分光器名称")
    onuMingCheng: Optional[str] = Field(default=None, description="ONU名称")

class FTTRCheckResponse(BaseModel):
    rows: List[Dict[str, Any]]
    preview: List[Dict[str, Any]]
    fileId: str
    filename: str
    downloadUrl: str
    rowCount: int

class SQLExportResponse(BaseModel):
    fileId: str
    filename: str
    downloadUrl: str
    rowCount: int

class DatasetDiffResponse(BaseModel):
    datasetKey: str
    displayName: str
    targetTable: str
    totalRows: int
    addedCount: int
    updatedCount: int
    deletedCount: int
    fileId: str
    filename: str
    downloadUrl: str
    uniqueColumns: List[str]

# Utility functions
# SQL 清洗：移除 Markdown 代码块包装与冗余空白，确保可直接执行
def clean_sql_query(sql: str) -> str:
    """Clean SQL query by removing markdown code blocks and extra whitespace"""
    if not sql or not isinstance(sql, str):
        return ""
    
    cleaned_sql = sql.strip()
    
    # Remove markdown code blocks
    if cleaned_sql.startswith("```sql") or cleaned_sql.startswith("```"):
        cleaned_sql = re.sub(r'^```(sql)?\n?', '', cleaned_sql)
        cleaned_sql = re.sub(r'\n?```$', '', cleaned_sql)
    
    # Remove extra whitespace
    cleaned_sql = cleaned_sql.strip()
    cleaned_sql = re.sub(r'\n\s*\n', '\n', cleaned_sql)
    
    return cleaned_sql

# SQL 校验：仅允许只读查询（SELECT/WITH），屏蔽增删改等危险操作
def validate_sql_query(sql: str) -> Dict[str, Any]:
    """Validate SQL query for safety"""
    if not sql or not isinstance(sql, str):
        return {"isValid": False, "error": "SQL查询语句不能为空"}
    
    cleaned_sql = clean_sql_query(sql).lower()
    
    if not cleaned_sql:
        return {"isValid": False, "error": "SQL查询语句为空"}
    
    # Check for dangerous operations
    dangerous_operations = [
        "drop table", "drop database", "truncate", "delete from",
        "update ", "insert into", "alter table", "create table", "create database"
    ]
    
    for operation in dangerous_operations:
        if operation in cleaned_sql:
            return {
                "isValid": False,
                "error": f"不允许执行 {operation.upper()} 操作，仅支持查询操作"
            }
    
    # Check if it starts with SELECT or WITH
    if not cleaned_sql.startswith("select") and not cleaned_sql.startswith("with"):
        return {"isValid": False, "error": "仅支持 SELECT 查询语句"}
    
    return {"isValid": True}

# 结果序列化：将 datetime 等类型转换为可 JSON 化的形式
def serialize_db_result(data: Any) -> Any:
    """Serialize database results, handling datetime and other types"""
    if data is None:
        return data
    
    if isinstance(data, datetime):
        return data.isoformat()
    
    if isinstance(data, list):
        return [serialize_db_result(item) for item in data]
    
    if isinstance(data, dict):
        return {key: serialize_db_result(value) for key, value in data.items()}
    
    return data

# 执行查询（返回字典列表）：统一连接池获取/释放与结果序列化
async def execute_query_dicts(sql: str, params: Optional[List[Any]] = None) -> List[Dict[str, Any]]:
    """Execute a SQL query and return list of dict rows."""
    global db_pool
    if not db_pool:
        raise HTTPException(status_code=500, detail="数据库连接不可用")
    params = params or []
    async with db_pool.acquire() as connection:
        rows = await connection.fetch(sql, *params)
        result: List[Dict[str, Any]] = []
        for r in rows:
            result.append(serialize_db_result(dict(r)))
        return result

# 导出结果为 Excel：按首行列顺序写入，并登记到 file_uploads 便于下载
async def export_rows_to_excel(rows: List[Dict[str, Any]], base_filename: str) -> Dict[str, str]:
    """Export rows to an Excel file under storage dir. Returns dict with id, filename, path."""
    if not rows:
        columns: List[str] = []
    else:
        # Preserve column order from first row
        columns = list(rows[0].keys())
    df = pd.DataFrame(rows, columns=columns)
    storage_dir = get_storage_dir()
    export_id = str(uuid.uuid4())
    ts = datetime.now().strftime("%Y%m%d%H%M%S")
    safe_base = re.sub(r"[^0-9a-zA-Z\u4e00-\u9fff_-]+", "_", base_filename).strip("_") or "export"
    filename = f"{safe_base}_{ts}.xlsx"
    path = os.path.join(storage_dir, filename)
    # Write Excel
    with pd.ExcelWriter(path, engine="openpyxl") as writer:
        df.to_excel(writer, sheet_name="结果", index=False)
    # Record in DB
    global db_pool
    if db_pool:
        try:
            async with db_pool.acquire() as conn:
                await ensure_migrations_tables(conn)
                await conn.execute(
                    "insert into file_uploads (id, filename, path, size_bytes, content_type, status) values ($1, $2, $3, $4, $5, 'generated')",
                    uuid.UUID(export_id), filename, path, os.path.getsize(path), "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
                )
        except Exception:
            pass
    return {"id": export_id, "filename": filename, "path": path}

# 简易意图识别（兜底）：基于关键词判断 OLT 统计 / FTTR 鉴别 / 未知
def recognize_task_from_text(text: str) -> str:
    """Return one of: 'OLT_STATISTICS', 'FTTR_CHECK', or 'UNKNOWN'"""
    t = (text or "").lower()
    if ("olt" in t) and ("统计" in text or "低效" in text or "数量" in text or "数" in text):
        return "OLT_STATISTICS"
    if ("fttr" in t) or ("分光" in text) or ("二级分光" in text) or ("onu" in t):
        return "FTTR_CHECK"
    return "UNKNOWN"

# 二级分光器名称抽取：支持引号内容、中文括号与包含 '/' 的模式
def extract_erji_fenguang_name(text: str) -> Optional[str]:
    """Extract 二级分光器名称 from free text.
    Improvements:
    - Support quoted names (single/double and Chinese quotes)
    - Allow full-width parentheses and more punctuation before '/'
    - Fall back to previous heuristics
    """
    if not text:
        return None
    t = text.strip()
    # Normalize Chinese quotes to ASCII quotes for easier matching
    t = t.replace(""", '"').replace(""", '"').replace("'", "'").replace("'", "'")
    # 1) Prefer content inside quotes that contains a '/'
    m_quote = re.search(r"""["']([^"'\n\r]+/[^"'\n\r]+)["']""", t)
    if m_quote:
        return m_quote.group(1).strip()
    # 2) Look for token containing '/' allowing Chinese full-width parentheses
    m = re.search(r'([\u4e00-\u9fffA-Za-z0-9_\-（）()·]+/[A-Za-z0-9_\-]+)', t)
    if m:
        return m.group(1).strip()
    # 3) Between keywords and decision words
    m2 = re.search(r"""(?:查询|判断|鉴别|查看|请帮我|请帮忙|二级分光器)\s*["']?(.+?)["']?\s*(?:能否|是否|能开通|能不能|可否|开通|fttr|FTTR)""", t)
    if m2:
        candidate = m2.group(1).strip()
        return candidate if candidate else None
    return None

# ONU 名称抽取：优先匹配“ONU用户/ONU”后引号内的名称
def extract_onu_name(text: str) -> Optional[str]:
    """Extract ONU用户名称 from free text.
    - Prefer names inside quotes near keywords like ONU/ONU用户
    - Support Chinese quotes and punctuation
    """
    if not text:
        return None
    t = text.strip()
    # Normalize quotes
    t = t.replace(""", '"').replace(""", '"').replace("'", "'").replace("'", "'")
    # 1) ONU用户 'xxx' or "xxx"
    m1 = re.search(r"""(?:ONU用户|onu用户|ONU|onu)\s*["']([^"'\n\r]+)["']""", t)
    if m1:
        return m1.group(1).strip()
    # 2) Generic quoted content when text mentions onu/ONU
    if re.search(r"\bonu\b", t, flags=re.IGNORECASE):
        m2 = re.search(r"""["']([^"'\n\r]+)["']""", t)
        if m2:
            return m2.group(1).strip()
    return None

OLT_SQL_TEMPLATE = (
    "with OLT_and_OLT_yonghu as\n"
    "(\n"
    "    SELECT\n"
    "    distinct\n"
    "    suo_shu_qu_xian,\n"
    "    suo_shu_ji_fang_zi_yuan_dian,\n"
    "    COUNT(*) OVER (PARTITION BY suo_shu_ji_fang_zi_yuan_dian) AS ji_fang_count,\n"
    "    SUM(CAST(onu_shu_liang as integer)) OVER (PARTITION BY suo_shu_ji_fang_zi_yuan_dian) AS yonghu_liang\n"
    "FROM\n"
    "    \"ziguan_olt_data\"\n"
    ")\n\n"
    "-- 低效\n"
    "SELECT\n"
    "    *,\n"
    "    CASE\n"
    "        WHEN ji_fang_count - CEILING(COALESCE(yonghu_liang, 0) / 4500.0) < 0 THEN 0\n"
    "        ELSE ji_fang_count - CEILING(COALESCE(yonghu_liang, 0) / 4500.0)\n"
    "    END AS dixiao_OLT_taishu\n"
    "FROM\n"
    "    OLT_and_OLT_yonghu;"
)

FTTR_FGQ_SQL_TEMPLATE = (
    "select A.fen_guang_qi_ming_cheng as erji_fen_guang,\n"
    "       A.fen_guang_qi_ji_bie,\n"
    "       B.fen_guang_qi_ming_cheng as yiji_fen_guang,\n"
    "       B.shang_lian_she_bei as OLT_mingcheng,\n"
    "       B.shang_lian_she_bei_zhu_yong_duan_kou as OLT_PON_kou,\n"
    "       E.suo_shu_ji_fang_zi_yuan_dian as jifang,\n"
    "       A.shang_lian_she_bei_zhu_yong_duan_kou,\n"
    "       A.shang_lian_she_bei_zhu_yong_duan_kou ~ 'CG' as support_open_FTTR\n"
    "from \"ziguan_fenguangqi\" A\n"
    "left join \"ziguan_fenguangqi\" B\n"
    "    on B.fen_guang_qi_ming_cheng = A.shang_lian_fen_guang_qi\n"
    "left join \"ziguan_olt_data\" E on E.olt_ming_cheng = B.shang_lian_she_bei\n"
    "where A.fen_guang_qi_ji_bie = '二级分光'"
)

FTTR_ONU_SQL_TEMPLATE = (
    "select C.onu_ming_cheng,\n"
    "       A.fen_guang_qi_ming_cheng as erji_fen_guang,\n"
    "       A.fen_guang_qi_ji_bie,\n"
    "       B.fen_guang_qi_ming_cheng as yiji_fen_guang,\n"
    "       B.shang_lian_she_bei as OLT_mingcheng,\n"
    "       B.shang_lian_she_bei_zhu_yong_duan_kou as OLT_PON_kou,\n"
    "       E.suo_shu_ji_fang_zi_yuan_dian as jifang,\n"
    "       A.shang_lian_she_bei_zhu_yong_duan_kou,\n"
    "       A.shang_lian_she_bei_zhu_yong_duan_kou ~ 'CG' as fenguangqi_support_open_FTTR,\n"
    "       C.zhong_duan_lei_xing,\n"
    "       C.zhong_duan_lei_xing in ('V176-20', 'HN8145XR', 'HG3142F', 'ZXHN G7611 V2', 'V175', 'V173', 'UNF130Z') as single_ONU_support_fttr\n\n"
    "from  \"wangguan_ONU_zaixianqingdan\" C\n"
    "join \"ziguan_ONU_guangmao\" D on C.onu_ming_cheng = D.xin_zeng_onu\n"
    "left join     ziguan_fenguangqi A on A.fen_guang_qi_ming_cheng = D.jie_ru_she_bei_ming_cheng\n"
    "left join \"ziguan_fenguangqi\" B\n"
    "    on B.fen_guang_qi_ming_cheng = A.shang_lian_fen_guang_qi\n"
    "left join \"ziguan_olt_data\" E on E.olt_ming_cheng = B.shang_lian_she_bei\n"
    "where A.fen_guang_qi_ji_bie = '二级分光'\n"
    "  and C.yun_xing_zhuang_tai = '在线'\n"
)

# 基于 LLM 的意图识别构建 SQL 模板：优先返回可直接预览/流式输出的模板
async def build_sql_for_intent(text: str, provider: Optional[str] = None) -> Dict[str, Any]:
    """Use LLM to recognize intent and build default SQL template for preview/stream."""
    llm_res = await recognize_intent_via_llm(text, provider)
    tasks = llm_res.get("tasks", [])
    # Choose first task for SQL template preview
    if not tasks:
        return {"task": "UNKNOWN", "sql": "", "message": "未识别到任务"}
    task_type = tasks[0].get("type")
    params = tasks[0].get("params", {})
    if task_type == "OLT_STATISTICS":
        return {"task": task_type, "sql": OLT_SQL_TEMPLATE, "params": params}
    if task_type == "FTTR_CHECK":
        onu_name = params.get("onuMingCheng")
        erji = params.get("erjiFenGuang")
        if onu_name:
            sql = FTTR_ONU_SQL_TEMPLATE + " and C.onu_ming_cheng = '" + str(onu_name).replace("'", "''") + "'"
            return {"task": task_type, "sql": sql, "params": params}
        if erji:
            erji_escaped = str(erji).replace("'", "''")
            sql = (
                "select A.fen_guang_qi_ming_cheng as erji_fen_guang,\n"
                "       A.fen_guang_qi_ji_bie,\n"
                "       B.fen_guang_qi_ming_cheng as yiji_fen_guang,\n"
                "       B.shang_lian_she_bei as OLT_mingcheng,\n"
                "       B.shang_lian_she_bei_zhu_yong_duan_kou as OLT_PON_kou,\n"
                "       E.suo_shu_ji_fang_zi_yuan_dian as jifang,\n"
                "       A.shang_lian_she_bei_zhu_yong_duan_kou,\n"
                "       A.shang_lian_she_bei_zhu_yong_duan_kou ~ 'CG' as support_open_FTTR  \n"
                "from \"ziguan_fenguangqi\" A\n"
                "left join \"ziguan_fenguangqi\" B\n"
                "    on B.fen_guang_qi_ming_cheng = A.shang_lian_fen_guang_qi\n"
                "left join \"ziguan_olt_data\" E on E.olt_ming_cheng = B.shang_lian_she_bei\n"
                "where A.fen_guang_qi_ji_bie = '二级分光'\n"
                f"and A.fen_guang_qi_ming_cheng = '{erji_escaped}'"
            )
            return {"task": task_type, "sql": sql, "params": params}
        return {"task": task_type, "sql": FTTR_FGQ_SQL_TEMPLATE, "params": params, "alternative": FTTR_ONU_SQL_TEMPLATE}
    return {"task": "UNKNOWN", "sql": "", "message": "未识别到任务"}

# 文件存储目录：统一集中到项目内 electronic-industry-agent/files
def get_storage_dir() -> str:
    base_dir = os.path.dirname(__file__)
    storage = os.path.join(base_dir, "electronic-industry-agent", "files")
    os.makedirs(storage, exist_ok=True)
    return storage

DATASET_PRESETS: Dict[str, Dict[str, Any]] = CONFIG_DATASET_PRESETS or {}

# 数据集配置读取：从预设中取得显示名/目标表/唯一列配置
def get_dataset_config(dataset_key: str) -> Dict[str, Any]:
    preset = DATASET_PRESETS.get(dataset_key)
    if not preset:
        raise HTTPException(status_code=400, detail=f"未知的数据集标识: {dataset_key}")
    return {
        "dataset_key": dataset_key,
        "display_name": preset["display_name"],
        "target_table": preset["target_table"],
        "unique_columns": list(preset.get("unique_columns") or []),
    }

# 差异比对值标准化：空白归一为 None，便于去重与对齐
def normalize_value_for_diff(value: Any) -> Optional[str]:
    if value is None:
        return None
    s = str(value)
    s = s.strip()
    if s == "":
        return None
    return s

# 基于唯一键列生成键值元组：用于判定新增/更新/重复
def compute_key_tuple(row: Dict[str, Any], unique_columns: List[str]) -> Tuple[Any, ...]:
    parts: List[Any] = []
    for c in unique_columns:
        parts.append(normalize_value_for_diff(row.get(c)))
    return tuple(parts)

# 将行列表转为 DataFrame：保持首行列顺序，导出更稳定
def rows_to_df(rows: List[Dict[str, Any]]) -> pd.DataFrame:
    if not rows:
        return pd.DataFrame()
    # Preserve first-row keys order
    cols = list(rows[0].keys())
    return pd.DataFrame(rows, columns=cols)

# 导出差异 Excel：分别输出“新增/修改/删除”三个工作表，并登记到库
async def export_diffs_excel(added: List[Dict[str, Any]], updated_new: List[Dict[str, Any]], deleted: List[Dict[str, Any]], base_filename: str) -> Dict[str, str]:
    storage_dir = get_storage_dir()
    export_id = str(uuid.uuid4())
    ts = datetime.now().strftime("%Y%m%d%H%M%S")
    safe_base = re.sub(r"[^0-9a-zA-Z\u4e00-\u9fff_-]+", "_", base_filename).strip("_") or "diff"
    filename = f"{safe_base}_{ts}.xlsx"
    path = os.path.join(storage_dir, filename)
    with pd.ExcelWriter(path, engine="openpyxl") as writer:
        rows_to_df(added).to_excel(writer, sheet_name="新增", index=False)
        rows_to_df(updated_new).to_excel(writer, sheet_name="修改", index=False)
        rows_to_df(deleted).to_excel(writer, sheet_name="删除", index=False)
    # record in DB
    global db_pool
    if db_pool:
        try:
            async with db_pool.acquire() as conn:
                await ensure_migrations_tables(conn)
                await conn.execute(
                    "insert into file_uploads (id, filename, path, size_bytes, content_type, status) values ($1, $2, $3, $4, $5, 'generated')",
                    uuid.UUID(export_id), filename, path, os.path.getsize(path), "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
                )
        except Exception:
            pass
    return {"id": export_id, "filename": filename, "path": path}

# 表存在性检测：使用 to_regclass 判定是否存在
async def table_exists(connection: asyncpg.Connection, table_name: str) -> bool:
    q = "select to_regclass($1) is not null as exists"
    try:
        exists = await connection.fetchval(q, table_name)
        return bool(exists)
    except Exception:
        return False

# 拉取整表所有行（谨慎用于小表/快照场景），自动处理引号
async def fetch_all_rows(connection: asyncpg.Connection, table_name: str) -> List[Dict[str, Any]]:
    try:
        rows = await connection.fetch(f'select * from "{table_name}"')
        return [serialize_db_result(dict(r)) for r in rows]
    except Exception:
        # Try without quotes when input already safe
        try:
            rows = await connection.fetch(f'select * from {table_name}')
            return [serialize_db_result(dict(r)) for r in rows]
        except Exception:
            return []

# 通用表格解析：支持 CSV/Excel，自动探测表头与生成安全列名
async def parse_tabular_file_to_rows(file_path: str) -> List[Dict[str, Any]]:
    # Handle CSV or Excel (first sheet)
    suffix = pathlib.Path(file_path).suffix.lower()
    rows: List[Dict[str, Any]] = []
    # 根据扩展名选择解析逻辑：CSV 与 Excel 走不同分支，输出统一的字典行列表
    if suffix == ".csv":
        # 优先尝试多种常见编码以避免解码错误（utf-8/gbk/latin1 等）
        enc = _detect_csv_encoding(file_path)
        with open(file_path, mode="r", encoding=enc, newline="") as f_sync:
            reader_sync = csv.reader(f_sync)
            try:
                headers = next(reader_sync)
            except StopIteration:
                return []
            # 统一表头：去两端空白，缺失置空字符串，便于后续生成列名
            headers = [str(h).strip() if h is not None else "" for h in headers]
            # Build pinyin-based columns with underscores
            try:
                from pinyin_utils import to_pinyin_list as _to_pinyin_list_cols
                base_names = [h if (h and isinstance(h, str) and h.strip() != "") else f"c_{i}" for i, h in enumerate(headers)]
                computed_columns = _to_pinyin_list_cols(base_names)
                # 二次清洗：确保列名只包含安全字符并转为下划线风格
                computed_columns = [sanitize_identifier(c) for c in computed_columns]
            except Exception:
                print("failed to import CSV headers by pinyin")
                computed_columns = [sanitize_identifier(h or f"c_{i}") for i, h in enumerate(headers)]
            for row in reader_sync:
                # 单元格标准化：空白统一为 None，便于后续对比与入库
                normalized = [normalize_value_for_diff(cell) for cell in row]
                if all(v is None for v in normalized):
                    continue
                # align length
                if len(normalized) < len(computed_columns):
                    normalized += [None] * (len(computed_columns) - len(normalized))
                elif len(normalized) > len(computed_columns):
                    normalized = normalized[: len(computed_columns)]
                rows.append({computed_columns[i]: normalized[i] for i in range(len(computed_columns))})
        return rows
    # Excel
    # 延迟导入以减少无关依赖；只读取首个工作表的数据
    from openpyxl import load_workbook
    wb = load_workbook(filename=file_path, read_only=True, data_only=True)
    ws = wb.worksheets[0]
    def non_empty_count(row_vals):
        return sum(1 for v in row_vals if v is not None and str(v).strip() != "")
    def effective_width(row_vals):
        last_idx = -1
        for idx, v in enumerate(row_vals):
            if v is not None and str(v).strip() != "":
                last_idx = idx
        return last_idx + 1
    # 扫描前 50 行用于定位最可能的表头行，并估算有效列宽
    scanned = list(ws.iter_rows(min_row=1, max_row=50, max_col=1024, values_only=True))
    header_row_index = 1
    header_non_empty = 0
    for idx, row_vals in enumerate(scanned, start=1):
        cnt = non_empty_count(row_vals)
        if cnt > header_non_empty and (cnt >= 2 or header_non_empty == 0):
            header_row_index = idx
            header_non_empty = cnt
    header_row = scanned[header_row_index - 1] if scanned else []
    headers = [str(h).strip() if h is not None else "" for h in header_row]
    num_cols = effective_width(header_row)
    print("Length of Columns: {}, columns: {}".format(num_cols, header_row[:num_cols]))
    if num_cols == 0:
        next_index = header_row_index + 1
        next_rows = list(ws.iter_rows(min_row=next_index, max_row=next_index, max_col=1024, values_only=True))
        first_data = next_rows[0] if next_rows else []
        num_cols = effective_width(first_data)
        headers = [f"col_{i}" for i in range(num_cols)]
        data_start_row = next_index
    else:
        if len(headers) < num_cols:
            headers = headers + [f"col_{i}" for i in range(len(headers), num_cols)]
        elif len(headers) > num_cols:
            headers = headers[:num_cols]
        data_start_row = header_row_index + 1
    # Build pinyin-based columns with underscores
    try:
        from pinyin_utils import to_pinyin_list as _to_pinyin_list_cols
        base_names = [h if (h and isinstance(h, str) and h.strip() != "") else f"c_{i}" for i, h in enumerate(headers)]
        columns = _to_pinyin_list_cols(base_names)
        # 列名安全化：移除非法字符并统一为小写下划线风格
        columns = [sanitize_identifier(c) for c in columns]


    except Exception:
        columns = [sanitize_identifier(h or f"c_{i}") for i, h in enumerate(headers)]
    def normalize_row(row_tuple):
        # 将整行转换为统一格式：空白为 None，长度对齐到列数
        row_list = [normalize_value_for_diff(cell) for cell in row_tuple]
        if all(v is None for v in row_list):
            return None
        if len(row_list) < len(columns):
            row_list += [None] * (len(columns) - len(row_list))
        elif len(row_list) > len(columns):
            row_list = row_list[: len(columns)]
        return row_list

    print("Length of New Columns: {}, columns: {}".format(len(columns), columns))
    data_rows_iter = ws.iter_rows(min_row=data_start_row, max_col=len(columns), values_only=True)
    for row in data_rows_iter:
        nr = normalize_row(row)
        if nr is None:
            continue
        rows.append({columns[i]: nr[i] for i in range(len(columns))})
    try:
        wb.close()
    except Exception:
        pass
    return rows

# 查询现有列顺序：用于增量添加缺失列
async def get_existing_columns(connection: asyncpg.Connection, table_name: str) -> List[str]:
    try:
        rows = await connection.fetch(
            """
            select column_name
            from information_schema.columns
            where table_schema = 'public' and table_name = $1
            order by ordinal_position
            """,
            table_name,
        )
        return [r["column_name"] for r in rows]
    except Exception:
        return []

# 确保目标表存在且列齐：不存在则建表，存在则补齐缺失列
async def ensure_target_table(connection: asyncpg.Connection, table_name: str, columns: List[str]) -> None:
    if not columns:
        return
    cols_defs = ", ".join([f'"{c}" text' for c in columns])
    await connection.execute(f'create table if not exists "{table_name}" ({cols_defs});')
    # Add missing columns if table already exists
    existing = await get_existing_columns(connection, table_name)
    existing_set = set(existing)
    for c in columns:
        if c not in existing_set:
            try:
                await connection.execute(f'alter table "{table_name}" add column "{c}" text;')
            except Exception:
                pass

# 批量插入：按给定列顺序批量写入，提高导入效率
async def bulk_insert_rows(connection: asyncpg.Connection, table_name: str, columns: List[str], rows: List[Dict[str, Any]]) -> int:
    if not rows:
        return 0
    placeholders = ", ".join([f"${i+1}" for i in range(len(columns))])
    insert_sql = f'insert into "{table_name}" ({", ".join([f"\"{c}\"" for c in columns])}) values ({placeholders})'
    values_batches: List[List[Any]] = []
    for r in rows:
        values_batches.append([normalize_value_for_diff(r.get(c)) for c in columns])
    await connection.executemany(insert_sql, values_batches)
    return len(rows)

# 以唯一键为条件的更新：仅更新非键列，参数化防注入
async def update_row_by_keys(connection: asyncpg.Connection, table_name: str, all_columns: List[str], key_columns: List[str], row: Dict[str, Any]) -> None:
    set_columns = [c for c in all_columns if c not in key_columns]
    if not set_columns:
        return
    set_clause = ", ".join([f'"{c}" = ${i+1}' for i, c in enumerate(set_columns)])
    where_clause = " and ".join([f'"{kc}" = ${len(set_columns) + idx + 1}' for idx, kc in enumerate(key_columns)])
    sql = f'update "{table_name}" set {set_clause} where {where_clause}'
    params: List[Any] = [normalize_value_for_diff(row.get(c)) for c in set_columns] + [normalize_value_for_diff(row.get(k)) for k in key_columns]
    await connection.execute(sql, *params)

# 以唯一键为条件的删除：当前主要用于对比类流程的占位
async def delete_row_by_keys(connection: asyncpg.Connection, table_name: str, key_columns: List[str], row: Dict[str, Any]) -> None:
    if not key_columns:
        return
    where_clause = " and ".join([f'"{kc}" = ${idx + 1}' for idx, kc in enumerate(key_columns)])
    sql = f'delete from "{table_name}" where {where_clause}'
    params: List[Any] = [normalize_value_for_diff(row.get(k)) for k in key_columns]
    await connection.execute(sql, *params)

# 数据集对比上传：解析上传文件与库内唯一键集合比对，输出新增/更新并落库
@app.post("/api/datasets/{dataset_key}/diff-upload", response_model=DatasetDiffResponse)
async def dataset_diff_upload(dataset_key: str, file: UploadFile = File(...)):
    global db_pool
    if not db_pool:
        raise HTTPException(status_code=500, detail="数据库连接不可用")
    if not file:
        raise HTTPException(status_code=400, detail="缺少上传文件")
    try:
        # 第一步：读取数据集配置与确保元表存在
        async with db_pool.acquire() as conn:
            await ensure_migrations_tables(conn)
            cfg = get_dataset_config(dataset_key)
            display_name = cfg["display_name"]
            target_table = cfg["target_table"]
            unique_columns: List[str] = list(cfg.get("unique_columns") or [])

        storage_dir = get_storage_dir()
        upload_id = str(uuid.uuid4())
        safe_name = re.sub(r"[^0-9a-zA-Z\u4e00-\u9fff_.-]+", "_", file.filename)
        target_path = os.path.join(storage_dir, f"{upload_id}_{safe_name}")
        size_bytes = 0
        async with aiofiles.open(target_path, 'wb') as out:
            while True:
                chunk = await file.read(1024 * 1024 * 4)
                if not chunk:
                    break
                size_bytes += len(chunk)
                await out.write(chunk)
        # Defer logging until rows are parsed to log row count instead of bytes

        # Parse uploaded rows
        new_rows = await parse_tabular_file_to_rows(target_path)
        total_rows = len(new_rows)
        print(f"[diff-upload] key={dataset_key} name={file.filename} size={total_rows}")

        # 基于数据库中“唯一键列”的去重集合来判定是新增还是更新
        # 说明：与其把整表读出来再比对整行（既耗时也占内存），
        #      这里只从数据库中查询 unique_columns 的去重值集合（distinct），
        #      然后用上传数据里的对应唯一键值去命中该集合：命中=更新，不命中=新增。
        #      一般不会有“删除行”的需求，删除视为 0，不在这里处理。
        # 第二步：构建库内唯一键“存在性集合”，用以判定新增/更新
        async with db_pool.acquire() as conn2:
            # 确保目标表存在且列齐全（会根据上传文件的列进行补充）
            detected_columns: List[str] = list(new_rows[0].keys()) if new_rows else []
            await ensure_target_table(conn2, target_table, detected_columns)
            # 组装并执行 DISTINCT 查询，仅拉取唯一键列的组合值
            # 注意：unique_columns 在上方将被保证至少有 1 列（若配置缺失，则兜底为首列）
            existing_key_set: Set[Tuple[Any, ...]] = set()
            if unique_columns:
                cols_sql = ", ".join([f'"{c}"' for c in unique_columns])
                sql_distinct = f'select distinct {cols_sql} from "{target_table}"'
                rows = await conn2.fetch(sql_distinct)
                for rec in rows:
                    # 仅保留唯一键列，构造与 compute_key_tuple 相容的字典
                    key_row = {c: rec.get(c) for c in unique_columns}
                    k = compute_key_tuple(key_row, unique_columns)
                    existing_key_set.add(k)

        print("unique columns: {}".format(unique_columns))
        print("detected existing key: ".format(len(existing_key_set)))
        # Derive unique columns if not configured
        if not unique_columns:
            if new_rows:
                unique_columns = [list(new_rows[0].keys())[0]]
            else:
                unique_columns = []

        # 注意：不再构建旧数据整行索引 old_index，仅保留唯一键的存在性集合 existing_key_set

        # 将上传行划分为：新增/更新/重复（上传文件内部重复）。不再处理“删除”。
        added_rows: List[Dict[str, Any]] = []
        updated_new_rows: List[Dict[str, Any]] = []
        duplicate_count = 0
        # seen_keys 用于识别“同一批上传数据内部”的重复键，避免重复计算和重复写入
        seen_keys: set[Tuple[Any, ...]] = set()

        # 取消“操作/状态列”的解析与删除行为识别。业务约定：一般不会有删除行，这里始终视为 0。

        def _rows_equal_excluding_keys(a: Dict[str, Any], b: Dict[str, Any], key_cols: List[str]) -> bool:
            a_norm = {c: normalize_value_for_diff(v) for c, v in a.items() if c not in key_cols}
            b_norm = {c: normalize_value_for_diff(v) for c, v in b.items() if c not in key_cols}
            all_cols = set(a_norm.keys()) | set(b_norm.keys())
            for c in all_cols:
                if (a_norm.get(c) or None) != (b_norm.get(c) or None):
                    return False
            return True

        # 遍历上传行：
        # - 若同一键在本次上传中已出现过，则计入 duplicate_count，跳过
        # - 若该键存在于数据库的 existing_key_set，则归为“更新”
        # - 否则归为“新增”
        # 第三步：遍历上传数据，利用 existing_key_set 判定类别
        for r in new_rows:
            k = compute_key_tuple(r, unique_columns)
            if k in seen_keys:
                duplicate_count += 1
                continue
            seen_keys.add(k)

            if k in existing_key_set:
                updated_new_rows.append(r)
            else:
                added_rows.append(r)

        # 分类结果日志：仅包含新增、更新、（上传内部）重复；删除固定为 0
        print(f"[diff-upload][classified] rows={total_rows} add={len(added_rows)} update={len(updated_new_rows)} duplicate={duplicate_count}")

        # Do not print per-change logs; only log uploaded row count above

        # Write-back to DB
        # 第四步：回写数据库——新增批量插入，更新按唯一键参数化更新
        async with db_pool.acquire() as connw:
            await ensure_target_table(connw, target_table, list(new_rows[0].keys()) if new_rows else [])
            # 根据上面的分类：新增执行插入；已存在的执行更新；不处理删除
            cols = list(new_rows[0].keys()) if new_rows else await get_existing_columns(connw, target_table)
            for r in added_rows:
                await bulk_insert_rows(connw, target_table, cols, [r])
            for r in updated_new_rows:
                await update_row_by_keys(connw, target_table, cols, unique_columns, r)

        # Special post-processing for 家客业务信息表
        if dataset_key == "jiake_yewu_xinxi":
            # 1) Detect mismatches between jiake_yewu_xinxi (A) and wangguan_ONU_zaixianqingdan (B)
            async with db_pool.acquire() as connv:
                mismatch_rows = await connv.fetch(
                    """
                    select A.xin_zeng_onu,
                           A.olt_ming_cheng,
                           A.olt_duan_kou,
                           B.wang_yuan_ming_cheng,
                           B.cao_hao,
                           B.duan_kou_hao
                    from jiake_yewu_xinxi as A
                    left join "wangguan_ONU_zaixianqingdan" as B
                      on A.xin_zeng_onu = B.onu_ming_cheng
                    where A.olt_ming_cheng is distinct from B.wang_yuan_ming_cheng
                       or split_part(A.olt_duan_kou, '-', 1) is distinct from B.cao_hao
                       or split_part(A.olt_duan_kou, '-', -1) is distinct from B.duan_kou_hao
                    """
                )
                mismatches: List[Dict[str, Any]] = [dict(r) for r in mismatch_rows]

            # Prepare sheet1 rows with side-by-side fields and splits
            sheet1_rows: List[Dict[str, Any]] = []
            for r in mismatches:
                a_port = (r.get("olt_duan_kou") or "")
                parts = str(a_port).split("-") if a_port is not None else []
                a_first = parts[0] if len(parts) >= 1 else None
                a_mid = parts[1] if len(parts) >= 2 else None
                a_last = parts[-1] if len(parts) >= 1 else None
                sheet1_rows.append({
                    "xin_zeng_onu": r.get("xin_zeng_onu"),
                    "A.olt_ming_cheng": r.get("olt_ming_cheng"),
                    "B.wang_yuan_ming_cheng": r.get("wang_yuan_ming_cheng"),
                    "A.olt_duan_kou": r.get("olt_duan_kou"),
                    "A.cao_hao": a_first,
                    "B.cao_hao": r.get("cao_hao"),
                    "A.zhong_jian_kuai": a_mid,
                    "A.duan_kou_hao": a_last,
                    "B.duan_kou_hao": r.get("duan_kou_hao"),
                })

            # Snapshot full table before update for row-wise comparison (sorted by xin_zeng_onu)
            async with db_pool.acquire() as connb:
                before_rows_full = await fetch_all_rows(connb, "jiake_yewu_xinxi")
            def sort_by_onu(rows: List[Dict[str, Any]]) -> List[Dict[str, Any]]:
                try:
                    return sorted(rows, key=lambda r: (str(r.get("xin_zeng_onu") or "")))
                except Exception:
                    return rows
            before_rows_sorted = sort_by_onu(before_rows_full)

            # 2) Apply database corrections on the mismatched rows
            async with db_pool.acquire() as connu:
                await connu.execute(
                    """
                    update jiake_yewu_xinxi as A
                    set olt_ming_cheng = B.wang_yuan_ming_cheng,
                        olt_duan_kou = B.cao_hao || '-' || split_part(A.olt_duan_kou, '-', 2) || '-' || B.duan_kou_hao
                    from "wangguan_ONU_zaixianqingdan" as B
                    where A.xin_zeng_onu = B.onu_ming_cheng
                      and (
                        A.olt_ming_cheng is distinct from B.wang_yuan_ming_cheng
                        or split_part(A.olt_duan_kou, '-', 1) is distinct from B.cao_hao
                        or split_part(A.olt_duan_kou, '-', -1) is distinct from B.duan_kou_hao
                      )
                    """
                )

            # 3) Fetch modified table rows for sheet2
            async with db_pool.acquire() as connr:
                modified_rows_full = await fetch_all_rows(connr, "jiake_yewu_xinxi")
            modified_rows_sorted = sort_by_onu(modified_rows_full)
            # Align after-rows columns to the same order as before-rows to ensure 1-1 mapping
            before_cols: List[str] = list(before_rows_sorted[0].keys()) if before_rows_sorted else (list(modified_rows_sorted[0].keys()) if modified_rows_sorted else [])
            aligned_before_rows: List[Dict[str, Any]] = [{c: r.get(c) for c in before_cols} for r in before_rows_sorted]
            aligned_after_rows: List[Dict[str, Any]] = [{c: r.get(c) for c in before_cols} for r in modified_rows_sorted]

            # 4) Export a two-sheet Excel with highlights
            storage_dir = get_storage_dir()
            export_id = str(uuid.uuid4())
            ts = datetime.now().strftime("%Y%m%d%H%M%S")
            safe_base = re.sub(r"[^0-9a-zA-Z\u4e00-\u9fff_-]+", "_", f"{display_name}-修正结果").strip("_") or "jiake_fix"
            filename = f"{safe_base}_{ts}.xlsx"
            path = os.path.join(storage_dir, filename)

            # Write via pandas, then apply openpyxl formatting
            with pd.ExcelWriter(path, engine="openpyxl") as writer:
                # 错误行：按照指定表头中文名称输出
                df_err = rows_to_df(sheet1_rows)
                header_labels = {
                    "xin_zeng_onu": "新增ONU名称",
                    "A.olt_ming_cheng": "家客信息表OLT名称",
                    "B.wang_yuan_ming_cheng": "网管OLT名称",
                    "A.olt_duan_kou": "家客信息表OLT端口",
                    "A.cao_hao": "家客信息表OLT槽号",
                    "B.cao_hao": "网管OLT槽号",
                    "A.zhong_jian_kuai": "家客信息表OLT中间字段",
                    "A.duan_kou_hao": "家客信息表OLT端口号",
                    "B.duan_kou_hao": "网管OLT端口号",
                }
                if not df_err.empty:
                    df_err = df_err.rename(columns={k: v for k, v in header_labels.items() if k in df_err.columns})
                df_err.to_excel(writer, sheet_name="错误行", index=False)
                # Ensure 修改前/修改后 have identical columns and row order for 1-1 mapping
                pd.DataFrame(aligned_before_rows, columns=before_cols).to_excel(writer, sheet_name="修改前表", index=False)
                pd.DataFrame(aligned_after_rows, columns=before_cols).to_excel(writer, sheet_name="修改后表", index=False)

                wb = writer.book
                red_fill = PatternFill(start_color="FFFF9999", end_color="FFFF9999", fill_type="solid")

                # Highlight mismatched fields in sheet1（基于中文表头）
                ws1 = wb["错误行"]
                if sheet1_rows:
                    headers1 = [cell.value for cell in ws1[1]]
                    col_idx_map1 = {name: idx + 1 for idx, name in enumerate(headers1)}
                    pairs = [
                        ("家客信息表OLT名称", "网管OLT名称"),
                        ("家客信息表OLT槽号", "网管OLT槽号"),
                        ("家客信息表OLT端口号", "网管OLT端口号"),
                    ]
                    for row_idx in range(2, ws1.max_row + 1):
                        for left_name, right_name in pairs:
                            li = col_idx_map1.get(left_name)
                            ri = col_idx_map1.get(right_name)
                            if not li or not ri:
                                continue
                            lv = ws1.cell(row=row_idx, column=li).value
                            rv = ws1.cell(row=row_idx, column=ri).value
                            if (lv or None) != (rv or None):
                                ws1.cell(row=row_idx, column=li).fill = red_fill
                                ws1.cell(row=row_idx, column=ri).fill = red_fill

                # Highlight differences between 修改前表 and 修改后表 row-by-row across all columns
                ws_before = wb["修改前表"]
                ws_after = wb["修改后表"]
                if ws_before.max_row == ws_after.max_row and ws_before.max_column == ws_after.max_column:
                    headers_common = [cell.value for cell in ws_before[1]]
                    for row_idx in range(2, ws_before.max_row + 1):
                        for col_idx in range(1, ws_before.max_column + 1):
                            v_before = ws_before.cell(row=row_idx, column=col_idx).value
                            v_after = ws_after.cell(row=row_idx, column=col_idx).value
                            if (v_before or None) != (v_after or None):
                                ws_before.cell(row=row_idx, column=col_idx).fill = red_fill
                                ws_after.cell(row=row_idx, column=col_idx).fill = red_fill

            # Register generated file into DB
            try:
                async with db_pool.acquire() as connr2:
                    await ensure_migrations_tables(connr2)
                    await connr2.execute(
                        "insert into file_uploads (id, filename, path, size_bytes, content_type, status) values ($1, $2, $3, $4, $5, 'generated')",
                        uuid.UUID(export_id), filename, path, os.path.getsize(path), "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
                    )
            except Exception:
                pass

            export = {"id": export_id, "filename": filename, "path": path}
        else:
            # 删除统一视为 0，这里传入空列表以兼容导出函数签名
            export = await export_diffs_excel(added_rows, updated_new_rows, [], base_filename=f"{display_name}-差异")

        return {
            "datasetKey": dataset_key,
            "displayName": display_name,
            "targetTable": target_table,
            "totalRows": total_rows,
            "addedCount": len(added_rows),
            "updatedCount": len(updated_new_rows),
            # 业务约定：不处理删除，这里固定返回 0
            "deletedCount": 0,
            "fileId": export["id"],
            "filename": export["filename"],
            "downloadUrl": f"/api/files/download/{export['id']}",
            "uniqueColumns": unique_columns,
        }
    except HTTPException:
        raise
    except Exception as e:
        tb = traceback.format_exc()
        print(f"[diff-upload][error] {e.__class__.__name__}: {e}\n{tb}")
        raise HTTPException(status_code=500, detail={
            "message": f"处理对比上传失败: {str(e)}",
            "errorType": e.__class__.__name__,
            "traceback": tb,
        })

# SQL 标识符清洗：转为安全的下划线小写形式，避免非法字符
def sanitize_identifier(name: str) -> str:
    cleaned = re.sub(r"[^0-9a-zA-Z_]+", "_", name.strip())
    cleaned = re.sub(r"_+", "_", cleaned)
    cleaned = cleaned.strip("_")
    if not cleaned:
        cleaned = "col"
    if cleaned[0].isdigit():
        cleaned = f"c_{cleaned}"
    return cleaned.lower()

def _is_chinese_char(ch: str) -> bool:
    return '\u4e00' <= ch <= '\u9fff'

# 依据文件名生成可读表名：去除 UUID/时间戳，仅保留中文/字母/连字符并转拼音
def compute_pretty_table_name(file_path: str) -> str:
    """Derive a pinyin-based table name from filename by:
    - Removing leading UUID and underscore
    - Removing trailing 14-digit timestamp
    - Keeping only Chinese characters, hyphens, and letters
    - Converting to pinyin using the same utility as columns
    """
    stem = pathlib.Path(file_path).stem
    # Remove leading UUID followed by underscore
    stem = re.sub(r'^[0-9a-fA-F]{8}(?:-[0-9a-fA-F]{4}){3}-[0-9a-fA-F]{12}_', '', stem)
    # Remove trailing 14-digit timestamp
    stem = re.sub(r'\d{14}$', '', stem)
    # Remove specific unwanted words
    stem = stem.replace('副本', '')
    # Filter allowed characters: Chinese, letters, hyphen
    filtered_chars = []
    for ch in stem:
        if _is_chinese_char(ch) or ch.isalpha() or ch == '-':
            filtered_chars.append(ch)
    filtered = ''.join(filtered_chars)
    filtered = re.sub(r'-{2,}', '-', filtered).strip('-')
    if not filtered:
        filtered = 'dataset'
    # Pinyin conversion (lazy import similar to Excel path)
    try:
        from pinyin_utils import to_pinyin_list as _to_pinyin_list_name
        pinyin_res = _to_pinyin_list_name([filtered])
        name = pinyin_res[0] if pinyin_res else 'dataset'
    except Exception:
        # Fallback: basic sanitize preserving hyphens
        name = re.sub(r'[^a-zA-Z\-]+', '_', filtered).strip('_') or 'dataset'
    # Normalize underscores and lowercase; keep hyphens
    name = re.sub(r'_+', '_', name).strip('_').lower()
    return name or 'dataset'

# 确保内部元表存在：file_uploads 与 csv_metadata，用于文件登记与结构复用
async def ensure_migrations_tables(connection: asyncpg.Connection) -> None:
    await connection.execute(
        """
        create table if not exists file_uploads (
            id uuid primary key,
            filename text not null,
            path text not null,
            size_bytes bigint not null,
            content_type text,
            status text not null default 'uploaded',
            dataset_table text,
            rows_imported bigint default 0,
            created_at timestamp not null default now(),
            updated_at timestamp not null default now()
        );
        
        create table if not exists csv_metadata (
            id uuid primary key,
            header_signature text unique not null,
            headers jsonb not null,
            columns jsonb not null,
            table_name text not null,
            created_at timestamp not null default now(),
            updated_at timestamp not null default now()
        );
        """
    )

def _try_read_csv_headers(file_path: str, encoding: str):
    with open(file_path, mode="r", encoding=encoding, newline="") as f_sync:
        reader_sync = csv.reader(f_sync)
        headers = next(reader_sync)
        return headers

# CSV 编码探测：按常见编码列表尝试读取表头并回退
def _detect_csv_encoding(file_path: str) -> str:
    """Best-effort CSV encoding detection with sensible fallbacks."""
    candidate_encodings = [
        "utf-8",
        "utf-8-sig",
        "gb18030",
        "gbk",
        "gb2312",
        "big5",
        "latin1",
    ]
    for enc in candidate_encodings:
        try:
            _ = _try_read_csv_headers(file_path, enc)
            return enc
        except UnicodeDecodeError:
            continue
        except StopIteration:
            # Empty file
            return enc
        except Exception:
            continue
    # Last resort
    return "utf-8"

# CSV 后台导入：按表头结构复用/创建数据表，分批插入并登记状态
async def import_csv_background(upload_id: str, file_path: str) -> None:
    global db_pool
    if not db_pool:
        return
    try:
        # 标记导入状态并准备表结构/元数据
        async with db_pool.acquire() as conn:
            await ensure_migrations_tables(conn)
            await conn.execute(
                "update file_uploads set status='importing', updated_at=now() where id=$1",
                uuid.UUID(upload_id),
            )

            path_obj = pathlib.Path(file_path)

            # Detect encoding and read headers
            detected_encoding = _detect_csv_encoding(file_path)
            print(f"[csv] start id={upload_id} path={file_path} encoding={detected_encoding}")
            # 单次读取表头用于生成“结构签名”，后续可复用相同结构的数据表
            with open(file_path, mode="r", encoding=detected_encoding, newline="") as f_sync:
                reader_sync = csv.reader(f_sync)
                headers = next(reader_sync)

            # Normalize headers for signature
            normalized_headers = [
                (h.strip().lower() if isinstance(h, str) else str(h).strip().lower()) for h in headers
            ]
            signature_src = json.dumps(normalized_headers, ensure_ascii=False)
            header_signature = hashlib.sha256(signature_src.encode("utf-8")).hexdigest()
            print(f"[csv] headers={headers}")

            # Check metadata for existing table
            rec_meta = await conn.fetchrow(
                "select table_name, columns from csv_metadata where header_signature=$1",
                header_signature,
            )
            if rec_meta:
                table_name = rec_meta["table_name"]
                columns = rec_meta["columns"]
                if not isinstance(columns, list):
                    try:
                        columns = json.loads(columns)
                    except Exception:
                        columns = []
                # Ensure table exists with expected columns
                column_defs_existing = ", ".join([f'"{c}" text' for c in columns])
                await conn.execute(f'create table if not exists "{table_name}" ({column_defs_existing});')
                # 同结构覆盖导入：截断旧数据，避免重复
                await conn.execute(f'truncate table "{table_name}"')
                print(f"[csv] reuse table={table_name} (truncate)")
            else:
                # Create new table and record metadata
                columns = [sanitize_identifier(h or f"col_{i}") for i, h in enumerate(headers)]
                base_name = compute_pretty_table_name(file_path)
                table_name = f"{base_name}_{header_signature[:8]}"
                column_defs = ", ".join([f'"{c}" text' for c in columns])
                await conn.execute(f'create table if not exists "{table_name}" ({column_defs});')
                await conn.execute(
                    """
                    insert into csv_metadata (id, header_signature, headers, columns, table_name)
                    values ($1::uuid, $2, $3::jsonb, $4::jsonb, $5)
                    on conflict (header_signature) do update
                    set headers = excluded.headers,
                        columns = excluded.columns,
                        table_name = excluded.table_name,
                        updated_at = now()
                    """,
                    uuid.uuid4(),
                    header_signature,
                    json.dumps(headers, ensure_ascii=False),
                    json.dumps(columns, ensure_ascii=False),
                    table_name,
                )
                print(f"[csv] create table={table_name} columns={columns}")

            rows_imported = 0
            batch_size = 1000
            insert_sql = f'insert into "{table_name}" ({", ".join([f"\"{c}\"" for c in columns])}) values ({", ".join([f"${i+1}" for i in range(len(columns))])})'

            # 将 CSV 数据分批（1000 行）插入，控制事务体量与内存
            def chunked(iterable, size):
                chunk: List[List[Optional[str]]] = []
                for row in iterable:
                    chunk.append(row)
                    if len(chunk) >= size:
                        yield chunk
                        chunk = []
                if chunk:
                    yield chunk

            # 二次读取文件（跳过表头），对齐列数并将空字符串规范为 NULL
            with open(file_path, mode="r", encoding=detected_encoding, newline="") as f_sync:
                reader_sync = csv.reader(f_sync)
                next(reader_sync)
                for batch in chunked(reader_sync, batch_size):
                    values_list = []
                    for row in batch:
                        row = [(cell if cell != '' else None) for cell in row]
                        if len(row) < len(columns):
                            row += [None] * (len(columns) - len(row))
                        elif len(row) > len(columns):
                            row = row[:len(columns)]
                        values_list.append(row)
                    await conn.executemany(insert_sql, values_list)
                    rows_imported += len(values_list)

            await conn.execute(
                "update file_uploads set status='imported', dataset_table=$2, rows_imported=$3, updated_at=now() where id=$1",
                uuid.UUID(upload_id),
                table_name,
                rows_imported,
            )
            print(f"[csv] done id={upload_id} table={table_name} rows={rows_imported}")
    except Exception as e:
        try:
            async with db_pool.acquire() as conn:
                await conn.execute(
                    "update file_uploads set status='failed', updated_at=now() where id=$1",
                    uuid.UUID(upload_id),
                )
        except Exception:
            pass
        print(f"CSV import failed: {e}")

# Excel 后台导入：智能识别表头行与有效列宽，转安全列名后批量写入
async def import_excel_background(upload_id: str, file_path: str) -> None:
    """Import first sheet of an Excel file as text columns."""
    global db_pool
    if not db_pool:
        return
    try:
        # 标记导入状态，随后构建/复用数据表结构
        async with db_pool.acquire() as conn:
            await ensure_migrations_tables(conn)
            await conn.execute(
                "update file_uploads set status='importing', updated_at=now() where id=$1",
                uuid.UUID(upload_id),
            )

            path_obj = pathlib.Path(file_path)

            # Lazy import to avoid hard dependency unless used
            from openpyxl import load_workbook
            print(f"[xlsx] start id={upload_id} path={file_path}")
            wb = load_workbook(filename=file_path, read_only=True, data_only=True)
            ws = wb.worksheets[0]

            # In read_only mode, worksheet dimensions can default to 'A1' until calculated.
            # Force dimension calculation to ensure iter_rows and max_row/max_column are correct.
            try:
                dimension_str = ws.calculate_dimension(force=True)
                print(f"[xlsx] dimension={dimension_str} max_row={ws.max_row} max_col={ws.max_column}")
            except Exception as e:
                print(f"[xlsx] warn: calculate_dimension failed: {e}")

            # Detect header row by scanning first N rows for max non-empty cells
            def non_empty_count(row_vals):
                return sum(1 for v in row_vals if v is not None and str(v).strip() != "")
            def effective_width(row_vals):
                last_idx = -1
                for idx, v in enumerate(row_vals):
                    if v is not None and str(v).strip() != "":
                        last_idx = idx
                return last_idx + 1
            # For large files, some generators write an incorrect dimension (e.g., A1:A1).
            # Avoid relying on ws.max_row/ws.max_column: read a wide column range and trim trailing empties.
            # 限定扫描范围，兼顾性能与鲁棒性
            max_scan_rows = 50
            scanned = list(ws.iter_rows(min_row=1, max_row=max_scan_rows, max_col=1024, values_only=True))

            header_row_index = 1
            header_non_empty = 0
            for idx, row_vals in enumerate(scanned, start=1):
                cnt = non_empty_count(row_vals)
                if cnt > header_non_empty and (cnt >= 2 or header_non_empty == 0):
                    header_row_index = idx
                    header_non_empty = cnt

            header_row = scanned[header_row_index - 1] if scanned else []
            headers = [str(h).strip() if h is not None else "" for h in header_row]
            # Determine the effective number of columns based on non-empty tail trimming
            num_cols = effective_width(header_row)
            print(f"[xlsx] sheet={ws.title} header_row={header_row_index} non_empty={header_non_empty}")
            print(f"[xlsx] headers={headers}")

            if num_cols == 0:
                next_index = header_row_index + 1
                next_rows = list(ws.iter_rows(min_row=next_index, max_row=next_index, max_col=1024, values_only=True))
                first_data = next_rows[0] if next_rows else []
                num_cols = effective_width(first_data)
                headers = [f"col_{i}" for i in range(num_cols)]
                data_start_row = next_index
            else:
                # Ensure headers length matches effective width
                if len(headers) < num_cols:
                    headers = headers + [f"col_{i}" for i in range(len(headers), num_cols)]
                elif len(headers) > num_cols:
                    headers = headers[:num_cols]
                data_start_row = header_row_index + 1

            # Build column names using Pinyin conversion utility
            try:
                from pinyin_utils import to_pinyin_list as _to_pinyin_list
            except Exception as e:
                print(f"[xlsx] warn: failed to import to_pinyin_list: {e}")
                def _to_pinyin_list(words, exclude_chars=['%']):
                    # Fallback: sanitize to a safe identifier if pinyin isn't available
                    return [sanitize_identifier(w) for w in words]

            base_names = [h if (h and isinstance(h, str) and h.strip() != "") else f"col_{i}" for i, h in enumerate(headers)]
            computed_columns = _to_pinyin_list(base_names)
            # Ensure final safety for SQL identifiers
            computed_columns = [sanitize_identifier(c) for c in computed_columns]

            # Build a stable header signature to deduplicate by structure
            normalized_headers = [
                (h.strip().lower() if isinstance(h, str) else str(h).strip().lower()) for h in headers
            ]
            signature_src = json.dumps(normalized_headers, ensure_ascii=False)
            header_signature = hashlib.sha256(signature_src.encode("utf-8")).hexdigest()

            # Try find existing metadata to reuse table
            rec_meta = await conn.fetchrow(
                "select table_name, columns from csv_metadata where header_signature=$1",
                header_signature,
            )
            if rec_meta:
                table_name = rec_meta["table_name"]
                columns = rec_meta["columns"]
                if not isinstance(columns, list):
                    try:
                        columns = json.loads(columns)
                    except Exception:
                        columns = []
                column_defs_existing = ", ".join([f'"{c}" text' for c in columns])
                await conn.execute(f'create table if not exists "{table_name}" ({column_defs_existing});')
                await conn.execute(f'truncate table "{table_name}"')
                print(f"[xlsx] reuse table={table_name} (truncate)")
            else:
                columns = computed_columns
                base_name = compute_pretty_table_name(file_path)
                table_name = f"{base_name}_{header_signature[:8]}"
                column_defs = ", ".join([f'"{c}" text' for c in columns])
                await conn.execute(f'create table if not exists "{table_name}" ({column_defs});')
                await conn.execute(
                    "insert into csv_metadata (id, header_signature, headers, columns, table_name) values ($1::uuid, $2, $3::jsonb, $4::jsonb, $5)",
                    uuid.uuid4(),
                    header_signature,
                    json.dumps(headers, ensure_ascii=False),
                    json.dumps(columns, ensure_ascii=False),
                    table_name,
                )
                print(f"[xlsx] create table={table_name} columns={columns}")

            insert_sql = f'insert into "{table_name}" ({", ".join([f"\"{c}\"" for c in columns])}) values ({", ".join([f"${i+1}" for i in range(len(columns))])})'

            # 以生成器形式按批次聚合行，避免一次性加载全部数据
            def chunked_rows(iterator, size):
                chunk: List[List[Optional[str]]] = []
                for row in iterator:
                    chunk.append(row)
                    if len(chunk) >= size:
                        yield chunk
                        chunk = []
                if chunk:
                    yield chunk

            rows_imported = 0
            batch_size = 1000

            # 逐行规范化：转字符串去空白，空值置 None，并裁剪/填充到固定列数
            def normalize_row(row_tuple):
                row_list = [
                    (str(cell).strip() if cell is not None and str(cell).strip() != "" else None)
                    for cell in row_tuple
                ]
                if all(v is None for v in row_list):
                    return None
                if len(row_list) < len(columns):
                    row_list += [None] * (len(columns) - len(row_list))
                elif len(row_list) > len(columns):
                    row_list = row_list[:len(columns)]
                return row_list

            # Iterate rows with explicit max_col to bypass incorrect worksheet dimensions
            data_rows_iter = ws.iter_rows(min_row=data_start_row, max_col=len(columns), values_only=True)
            normalized_iter = (normalize_row(r) for r in data_rows_iter)
            filtered_iter = (r for r in normalized_iter if r is not None)

            preview = []
            for _ in range(2):
                try:
                    nxt = next(filtered_iter)
                    preview.append(nxt)
                except StopIteration:
                    break
            if preview:
                print(f"[xlsx] sample_row_1={preview[0]}")
                if len(preview) > 1:
                    print(f"[xlsx] sample_row_2={preview[1]}")

            # 使用独立连接批量写入，优先插入两行预览样本便于快速确认
            async with db_pool.acquire() as conn2:
                if preview:
                    await conn2.executemany(insert_sql, preview)
                    rows_imported += len(preview)
                for batch in chunked_rows(filtered_iter, batch_size):
                    await conn2.executemany(insert_sql, batch)
                    rows_imported += len(batch)

            await conn.execute(
                "update file_uploads set status='imported', dataset_table=$2, rows_imported=$3, updated_at=now() where id=$1",
                uuid.UUID(upload_id),
                table_name,
                rows_imported,
            )
            print(f"[xlsx] done id={upload_id} table={table_name} rows={rows_imported}")
            try:
                wb.close()
            except Exception:
                pass
    except Exception as e:
        try:
            async with db_pool.acquire() as conn:
                await conn.execute(
                    "update file_uploads set status='failed', updated_at=now() where id=$1",
                    uuid.UUID(upload_id),
                )
        except Exception:
            pass
        print(f"Excel import failed: {e}")


# LLM clients
# OpenAI 客户端：封装超时与重试，提供标准与流式输出
class OpenAIClient:
    """OpenAI API client"""
    
    def __init__(self):
        if not OPENAI_API_KEY:
            raise ValueError("未配置OpenAI API密钥。请在环境变量中设置OPENAI_API_KEY。")
    
    async def generate(self, prompt: str) -> str:
        """Generate response from OpenAI API with timeout and retries."""
        timeout = httpx.Timeout(LLM_HTTP_TIMEOUT_SECONDS)
        for attempt in range(LLM_MAX_RETRIES + 1):
            try:
                async with httpx.AsyncClient(timeout=timeout) as client:
                    response = await client.post(
                        OPENAI_API_ENDPOINT,
                        headers={
                            "Content-Type": "application/json",
                            "Authorization": f"Bearer {OPENAI_API_KEY}"
                        },
                        json={
                            "model": OPENAI_MODEL,
                            "messages": [{"role": "user", "content": prompt}],
                            "temperature": 0.1
                        }
                    )
                    response.raise_for_status()
                    data = response.json()
                    return data["choices"][0]["message"]["content"].strip()
            except (httpx.ReadTimeout, httpx.ConnectTimeout):
                if attempt < LLM_MAX_RETRIES:
                    await asyncio.sleep(LLM_BACKOFF_BASE * (2 ** attempt))
                    continue
                raise HTTPException(status_code=500, detail="OpenAI API Error: Timeout")
            except httpx.HTTPStatusError as he:
                status = he.response.status_code if he.response else ""
                body = (he.response.text if he.response else "")[:500]
                raise HTTPException(status_code=500, detail=f"OpenAI API Error: status={status}, body={body}")
            except httpx.RequestError as re_err:
                raise HTTPException(status_code=500, detail=f"OpenAI API Error: {re_err.__class__.__name__}: {str(re_err)}")
            except Exception as e:
                raise HTTPException(status_code=500, detail=f"OpenAI API Error: {e.__class__.__name__}: {str(e)}")
    
    async def generate_stream(self, prompt: str) -> AsyncGenerator[str, None]:
        """Generate streaming response from OpenAI API"""
        timeout = httpx.Timeout(LLM_HTTP_TIMEOUT_SECONDS)
        async with httpx.AsyncClient(timeout=timeout) as client:
            try:
                async with client.stream(
                    "POST",
                    OPENAI_API_ENDPOINT,
                    headers={
                        "Content-Type": "application/json",
                        "Authorization": f"Bearer {OPENAI_API_KEY}"
                    },
                    json={
                        "model": OPENAI_MODEL,
                        "messages": [{"role": "user", "content": prompt}],
                        "temperature": 0.1,
                        "stream": True
                    }
                ) as response:
                    response.raise_for_status()
                    
                    buffer = ""
                    async for chunk in response.aiter_text():
                        buffer += chunk
                        lines = buffer.split("\n")
                        buffer = lines.pop()
                        
                        for line in lines:
                            if line.startswith("data: "):
                                data = line[6:]
                                if data == "[DONE]":
                                    return
                                
                                try:
                                    parsed = json.loads(data)
                                    content = parsed.get("choices", [{}])[0].get("delta", {}).get("content")
                                    if content:
                                        yield content
                                except json.JSONDecodeError:
                                    continue
            except Exception as e:
                # Streaming errors
                if isinstance(e, httpx.HTTPStatusError):
                    he = e
                    status = he.response.status_code if he.response else ""
                    body = (he.response.text if he.response else "")[:500]
                    raise HTTPException(status_code=500, detail=f"OpenAI API Error: status={status}, body={body}")
                if isinstance(e, httpx.RequestError):
                    raise HTTPException(status_code=500, detail=f"OpenAI API Error: {e.__class__.__name__}: {str(e)}")
                raise HTTPException(status_code=500, detail=f"OpenAI API Error: {e.__class__.__name__}: {str(e)}")

# Gemini 客户端：简单封装，提供“伪流式”分片输出
class GeminiClient:
    """Gemini API client"""
    
    def __init__(self):
        if not GEMINI_API_KEY:
            raise ValueError("未配置Gemini API密钥。请在环境变量中设置GEMINI_API_KEY。")
    
    async def generate(self, prompt: str) -> str:
        """Generate response from Gemini API with timeout and retries."""
        timeout = httpx.Timeout(LLM_HTTP_TIMEOUT_SECONDS)
        for attempt in range(LLM_MAX_RETRIES + 1):
            try:
                async with httpx.AsyncClient(timeout=timeout) as client:
                    response = await client.post(
                        f"{GEMINI_ENDPOINT}?key={GEMINI_API_KEY}",
                        headers={"Content-Type": "application/json"},
                        json={
                            "contents": [{"parts": [{"text": prompt}]}],
                            "generationConfig": {"temperature": 0.1}
                        }
                    )
                    response.raise_for_status()
                    data = response.json()
                    return data["candidates"][0]["content"]["parts"][0]["text"].strip()
            except (httpx.ReadTimeout, httpx.ConnectTimeout):
                if attempt < LLM_MAX_RETRIES:
                    await asyncio.sleep(LLM_BACKOFF_BASE * (2 ** attempt))
                    continue
                raise HTTPException(status_code=500, detail="Gemini API Error: Timeout")
            except httpx.HTTPStatusError as he:
                status = he.response.status_code if he.response else ""
                body = (he.response.text if he.response else "")[:500]
                raise HTTPException(status_code=500, detail=f"Gemini API Error: status={status}, body={body}")
            except httpx.RequestError as re_err:
                raise HTTPException(status_code=500, detail=f"Gemini API Error: {re_err.__class__.__name__}: {str(re_err)}")
            except Exception as e:
                raise HTTPException(status_code=500, detail=f"Gemini API Error: {e.__class__.__name__}: {str(e)}")
    
    async def generate_stream(self, prompt: str) -> AsyncGenerator[str, None]:
        """Generate streaming response from Gemini API (simulated)"""
        # Note: Gemini doesn't have true streaming, so we simulate it
        content = await self.generate(prompt)
        
        # Split content into chunks and yield with small delays
        chunks = list(content)
        for chunk in chunks:
            yield chunk
            await asyncio.sleep(0.01)  # Small delay to simulate streaming

# DeepSeek 客户端：兼容 OpenAI Chat Completions 接口与流式输出
class DeepSeekClient:
    """DeepSeek API client (OpenAI-compatible chat completions)."""
    
    def __init__(self):
        if not DEEPSEEK_API_KEY:
            raise ValueError("未配置DeepSeek API密钥。请在环境变量中设置DEEPSEEK_API_KEY。")
    
    async def generate(self, prompt: str) -> str:
        timeout = httpx.Timeout(LLM_HTTP_TIMEOUT_SECONDS)
        for attempt in range(LLM_MAX_RETRIES + 1):
            try:
                async with httpx.AsyncClient(timeout=timeout) as client:
                    response = await client.post(
                        DEEPSEEK_API_ENDPOINT,
                        headers={
                            "Content-Type": "application/json",
                            "Authorization": f"Bearer {DEEPSEEK_API_KEY}"
                        },
                        json={
                            "model": DEEPSEEK_MODEL,
                            "messages": [{"role": "user", "content": prompt}],
                            "temperature": 0.1
                        }
                    )
                    response.raise_for_status()
                    data = response.json()
                    return data["choices"][0]["message"]["content"].strip()
            except (httpx.ReadTimeout, httpx.ConnectTimeout):
                if attempt < LLM_MAX_RETRIES:
                    await asyncio.sleep(LLM_BACKOFF_BASE * (2 ** attempt))
                    continue
                raise HTTPException(status_code=500, detail="DeepSeek API Error: Timeout")
            except httpx.HTTPStatusError as he:
                status = he.response.status_code if he.response else ""
                body = (he.response.text if he.response else "")[:500]
                raise HTTPException(status_code=500, detail=f"DeepSeek API Error: status={status}, body={body}")
            except httpx.RequestError as re_err:
                raise HTTPException(status_code=500, detail=f"DeepSeek API Error: {re_err.__class__.__name__}: {str(re_err)}")
            except Exception as e:
                raise HTTPException(status_code=500, detail=f"DeepSeek API Error: {e.__class__.__name__}: {str(e)}")
    
    async def generate_stream(self, prompt: str) -> AsyncGenerator[str, None]:
        timeout = httpx.Timeout(LLM_HTTP_TIMEOUT_SECONDS)
        async with httpx.AsyncClient(timeout=timeout) as client:
            try:
                async with client.stream(
                    "POST",
                    DEEPSEEK_API_ENDPOINT,
                    headers={
                        "Content-Type": "application/json",
                        "Authorization": f"Bearer {DEEPSEEK_API_KEY}"
                    },
                    json={
                        "model": DEEPSEEK_MODEL,
                        "messages": [{"role": "user", "content": prompt}],
                        "temperature": 0.1,
                        "stream": True
                    }
                ) as response:
                    response.raise_for_status()
                    
                    buffer = ""
                    async for chunk in response.aiter_text():
                        buffer += chunk
                        lines = buffer.split("\n")
                        buffer = lines.pop()
                        
                        for line in lines:
                            if line.startswith("data: "):
                                data = line[6:]
                                if data == "[DONE]":
                                    return
                                try:
                                    parsed = json.loads(data)
                                    content = parsed.get("choices", [{}])[0].get("delta", {}).get("content")
                                    if content:
                                        yield content
                                except json.JSONDecodeError:
                                    continue
            except Exception as e:
                if isinstance(e, httpx.HTTPStatusError):
                    he = e
                    status = he.response.status_code if he.response else ""
                    body = (he.response.text if he.response else "")[:500]
                    raise HTTPException(status_code=500, detail=f"DeepSeek API Error: status={status}, body={body}")
                if isinstance(e, httpx.RequestError):
                    raise HTTPException(status_code=500, detail=f"DeepSeek API Error: {e.__class__.__name__}: {str(e)}")
                raise HTTPException(status_code=500, detail=f"DeepSeek API Error: {e.__class__.__name__}: {str(e)}")

# API endpoints
# LLM 客户端工厂：按 provider 选择对应实现，默认 DeepSeek
def get_llm_client(provider: Optional[str] = None):
    name = (provider or LLM_PROVIDER or "deepseek").lower()
    if name == "openai":
        return OpenAIClient()
    if name == "gemini":
        return GeminiClient()
    return DeepSeekClient()

# 调用 LLM 做意图识别：强制 JSON 输出并做健壮性清洗
async def recognize_intent_via_llm(text: str, provider: Optional[str] = None) -> Dict[str, Any]:
    client = get_llm_client(provider)
    prompt = build_intent_prompt(text)
    content = await client.generate(prompt)
    try:
        # Some models may wrap code fences; strip if present
        cleaned = content.strip()
        if cleaned.startswith("```"):
            cleaned = re.sub(r'^```(json|JSON)?\n?', '', cleaned)
            cleaned = re.sub(r'\n?```$', '', cleaned)
        data = json.loads(cleaned)
        if not isinstance(data, dict):
            raise ValueError("LLM输出不是JSON对象")
        tasks = data.get("tasks", [])
        if not isinstance(tasks, list):
            tasks = []
        # Normalize
        norm_tasks: List[Dict[str, Any]] = []
        for t in tasks:
            ttype = (t or {}).get("type")
            params = (t or {}).get("params") or {}
            if ttype in ("OLT_STATISTICS", "FTTR_CHECK"):
                norm_tasks.append({"type": ttype, "params": params})
        return {"tasks": norm_tasks}
    except Exception as e:
        # Fallback: no tasks
        return {"tasks": []}

# 根路由：用于服务可用性与版本验证
@app.get("/")
async def root():
    """Root endpoint"""
    return {"message": "Electronic Industry Agent Backend", "version": "1.0.0"}

# 简单 LLM 接口：仅返回固定 SQL 模板（不做自由生成）
@app.post("/api/call-llm")
async def call_llm(request: LLMRequest):
    """Return fixed SQL templates based on intent recognition (no free-form generation)."""
    try:
        if not request.userInput:
            raise HTTPException(status_code=400, detail="缺少用户输入")
        intent_sql = await build_sql_for_intent(request.userInput, request.modelType)
        print(f"[LLM] input={request.userInput} -> intent={intent_sql.get('task')} sql_len={len(intent_sql.get('sql', ''))}")
        return {"sql": intent_sql.get("sql", "")}
    except HTTPException:
        # Preserve detailed upstream error
        raise
    except Exception as e:
        raise HTTPException(status_code=500, detail=f"LLM处理错误: {e.__class__.__name__}: {str(e)}")

# 流式 LLM 接口：先返回“思考”再按块推送 SQL 模板
@app.post("/api/call-llm-stream")
async def call_llm_stream(request: LLMStreamRequest):
    """Stream fixed SQL template chunks based on intent (UTF-8 JSON, unescaped)."""

    async def generate_stream():
        try:
            if not request.userInput:
                yield f"data: {json.dumps({'error': '缺少用户输入'}, ensure_ascii=False)}\n\n"
                return

            intent_sql = await build_sql_for_intent(request.userInput)
            task = intent_sql.get("task")
            sql_text = intent_sql.get("sql", "")
            print(f"[LLM-STREAM] input={request.userInput} -> intent={task} sql_len={len(sql_text)}")
            thinking_part = f"识别任务: {task}"
            params = intent_sql.get("params", {})

            # Send thinking first
            yield f"data: {json.dumps({'thinking': thinking_part, 'sql': '', 'isComplete': False, 'params': params}, ensure_ascii=False)}\n\n"

            # Stream SQL in chunks
            chunk_size = 200
            for i in range(0, len(sql_text), chunk_size):
                partial_sql = sql_text[: i + chunk_size]
                payload = {"thinking": thinking_part, "sql": partial_sql, "isComplete": False, "params": params}
                yield f"data: {json.dumps(payload, ensure_ascii=False)}\n\n"

            # Completion
            yield f"data: {json.dumps({'thinking': thinking_part, 'sql': sql_text, 'isComplete': True, 'params': params}, ensure_ascii=False)}\n\n"

        except HTTPException as he:
            # Surface detailed upstream error
            error_data = {"error": he.detail}
            yield f"data: {json.dumps(error_data, ensure_ascii=False)}\n\n"
        except Exception as e:
            # Generic error with type
            error_data = {"error": f"{e.__class__.__name__}: {str(e)}"}
            yield f"data: {json.dumps(error_data, ensure_ascii=False)}\n\n"

    return StreamingResponse(
        generate_stream(),
        media_type="text/event-stream",
        headers={
            "Cache-Control": "no-cache",
            "Connection": "keep-alive"
        }
    )

# 意图识别接口：返回 OLT 统计/FTTR 鉴别任务以及抽取的参数
@app.post("/api/intent/recognize", response_model=IntentResponse)
async def recognize_intent(request: IntentRequest):
    """LLM-based intent recognition for OLT统计 and FTTR鉴别."""
    text = (request.text or "").strip()
    llm_res = await recognize_intent_via_llm(text)
    tasks_data = llm_res.get("tasks", [])
    tasks: List[IntentTask] = []
    for t in tasks_data:
        tasks.append(IntentTask(type=t.get("type", "UNKNOWN"), params=t.get("params", {})))
    if not tasks:
        tasks = [IntentTask(type="OLT_STATISTICS", params={}), IntentTask(type="FTTR_CHECK", params={})]
    return {"tasks": tasks}

# 意图执行接口：自动识别后直接执行对应任务并返回完整结果
@app.post("/api/intent/execute")
async def execute_intent(request: IntentRequest):
    """Execute intent end-to-end: use LLM for intent, then run the appropriate task.
    Returns full task result so FTTR (ONU) will include the step-2 aggregation when needed.
    """
    text = (request.text or "").strip()
    llm_res = await recognize_intent_via_llm(text)
    tasks = llm_res.get("tasks", [])
    if not tasks:
        raise HTTPException(status_code=400, detail="未识别到任务")
    first = tasks[0]
    ttype = first.get("type")
    params = first.get("params", {})
    if ttype == "FTTR_CHECK":
        req = FTTRCheckRequest(
            erjiFenGuang=params.get("erjiFenGuang"),
            onuMingCheng=params.get("onuMingCheng"),
        )
        return await task_fttr_check(req)
    if ttype == "OLT_STATISTICS":
        return await task_olt_statistics()
    raise HTTPException(status_code=400, detail="未识别到可执行的任务类型")

# SQL 查询执行端点：清洗与校验 SQL，只允许只读查询，返回数据/实体信息/推荐
@app.post("/api/sql-query")
async def execute_sql_query(request: SQLQueryRequest):
    """Execute SQL query"""
    global db_pool
    
    if not db_pool:
        raise HTTPException(status_code=500, detail="数据库连接不可用")
    
    try:
        if not request.sql:
            raise HTTPException(status_code=400, detail="缺少SQL查询语句")
        
        # 第一步：清洗/校验 SQL，只允许只读查询（SELECT/WITH）
        cleaned_sql = clean_sql_query(request.sql)
        
        if not cleaned_sql:
            raise HTTPException(status_code=400, detail="SQL查询语句为空")
        
        validation = validate_sql_query(cleaned_sql)
        if not validation["isValid"]:
            raise HTTPException(status_code=400, detail=validation.get("error", "SQL查询语句无效"))
        
        # 第二步：执行查询并序列化结果
        async with db_pool.acquire() as connection:
            try:
                result = await connection.fetch(cleaned_sql)
                
                # Convert result to list of dictionaries
                data = []
                for row in result:
                    row_dict = dict(row)
                    data.append(serialize_db_result(row_dict))

                # 第三步：若未显式传入实体信息，尝试基于 SQL 文本推断实体类型与名称
                entity_type = request.entityType
                entity_name = request.entityName
                try:
                    lowered = cleaned_sql.lower()
                    if not entity_type:
                        if ("onu" in lowered) or ("光猫" in cleaned_sql):
                            entity_type = "ONU"
                        elif ("fen_guang_qi" in lowered) or ("分光器" in cleaned_sql):
                            entity_type = "分光器"
                    if not entity_name:
                        # Extract all quoted literals and choose the most likely candidate
                        candidates: list[str] = []
                        candidates += [m.strip() for m in re.findall(r"'(.*?)'", cleaned_sql)]
                        candidates += [m.strip() for m in re.findall(r"“(.*?)”", cleaned_sql)]
                        candidates += [m.strip() for m in re.findall(r"「(.*?)」", cleaned_sql)]
                        # Prefer those containing 'ONU' (case-insensitive), otherwise the longest one
                        best = None
                        for s in candidates:
                            if not best:
                                best = s
                                continue
                            if ("onu" in s.lower()) and ("onu" not in best.lower()):
                                best = s
                                continue
                            if len(s) > len(best):
                                best = s
                        if best:
                            entity_name = best
                except Exception:
                    pass

                # 第四步（可选）：若为 ONU 且不支持 CG 口，基于机房给出同机房推荐列表
                recommendations: List[Dict[str, Any]] = []
                try:
                    def _support_flag_js(rec: Dict[str, Any]) -> bool:
                        return bool(rec.get("fenguangqi_support_open_FTTR") or rec.get("fenguangqi_support_open_fttr") or rec.get("support_open_FTTR") or rec.get("support_open_fttr"))

                    if entity_type == "ONU" and data:
                        any_support = any(_support_flag_js(r) for r in data)
                        first = data[0]
                        jifang_val = first.get("jifang") if isinstance(first, dict) else None
                        if (not any_support) and jifang_val:
                            sql_step2 = (
                                "select\n"
                                "    A.fen_guang_qi_ming_cheng as erji_fen_guang,\n"
                                "    COUNT(C.onu_ming_cheng) as onu_count,\n"
                                "    A.fen_guang_qi_ji_bie,\n"
                                "    B.fen_guang_qi_ming_cheng as yiji_fen_guang,\n"
                                "    B.shang_lian_she_bei as OLT_mingcheng,\n"
                                "    B.shang_lian_she_bei_zhu_yong_duan_kou as OLT_PON_kou,\n"
                                "    E.suo_shu_ji_fang_zi_yuan_dian as jifang,\n"
                                "    A.shang_lian_she_bei_zhu_yong_duan_kou,\n"
                                "    A.shang_lian_she_bei_zhu_yong_duan_kou ~ 'CG' as fenguangqi_support_open_FTTR,\n"
                                "    STRING_AGG(C.zhong_duan_lei_xing, ', ') as zhong_duan_lei_xing_list,\n"
                                "    BOOL_OR(C.zhong_duan_lei_xing in ('V176-20', 'HN8145XR', 'HG3142F', 'ZXHN G7611 V2', 'V175', 'V173', 'UNF130Z')) as has_single_ONU_support_fttr\n\n"
                                "from \"wangguan_ONU_zaixianqingdan\" C\n"
                                "join \"ziguan_ONU_guangmao\" D on C.onu_ming_cheng = D.xin_zeng_onu\n"
                                "left join \"ziguan_fenguangqi\" A on A.fen_guang_qi_ming_cheng = D.jie_ru_she_bei_ming_cheng\n"
                                "left join \"ziguan_fenguangqi\" B\n"
                                "    on B.fen_guang_qi_ming_cheng = A.shang_lian_fen_guang_qi\n"
                                "left join \"ziguan_olt_data\" E on E.olt_ming_cheng = B.shang_lian_she_bei\n"
                                "where A.fen_guang_qi_ji_bie = '二级分光'\n"
                                "  and E.suo_shu_ji_fang_zi_yuan_dian = $1 and (A.shang_lian_she_bei_zhu_yong_duan_kou ~ 'CG') and C.zhong_duan_lei_xing not in ('V176-20', 'HN8145XR', 'HG3142F', 'ZXHN G7611 V2', 'V175', 'V173', 'UNF130Z')\n"
                                "  and C.yun_xing_zhuang_tai = '在线'\n"
                                "GROUP BY\n"
                                "    A.fen_guang_qi_ming_cheng,\n"
                                "    A.fen_guang_qi_ji_bie,\n"
                                "    B.fen_guang_qi_ming_cheng,\n"
                                "    B.shang_lian_she_bei,\n"
                                "    B.shang_lian_she_bei_zhu_yong_duan_kou,\n"
                                "    E.suo_shu_ji_fang_zi_yuan_dian,\n"
                                "    A.shang_lian_she_bei_zhu_yong_duan_kou\n"
                                "ORDER BY onu_count DESC;\n"
                            )
                            recommendations = await execute_query_dicts(sql_step2, [jifang_val])
                except Exception:
                    pass

                # 返回结构：data 为结果数组，同时附带 entityType/entityName 与 recommendations
                return {
                    "data": data,
                    "entityType": entity_type,
                    "entityName": entity_name,
                    "recommendations": recommendations,
                }
            
            except Exception as e:
                raise HTTPException(status_code=500, detail=f"SQL查询错误: {str(e)}")
    
    except HTTPException:
        raise
    except Exception as e:
        raise HTTPException(status_code=500, detail=f"处理请求时发生错误: {str(e)}")

# SQL 查询导出端点：执行查询后整表导出为 Excel 并提供下载
@app.post("/api/sql-query/export", response_model=SQLExportResponse)
async def export_sql_query(request: SQLQueryRequest):
    """Execute a SQL query and export full result to Excel, return download info."""
    global db_pool
    if not db_pool:
        raise HTTPException(status_code=500, detail="数据库连接不可用")
    if not request.sql:
        raise HTTPException(status_code=400, detail="缺少SQL查询语句")
    cleaned_sql = clean_sql_query(request.sql)
    validation = validate_sql_query(cleaned_sql)
    if not validation["isValid"]:
        raise HTTPException(status_code=400, detail=validation.get("error", "SQL查询语句无效"))
    print("[SQL-EXPORT] executing export for SQL len=", len(cleaned_sql))
    rows = await execute_query_dicts(cleaned_sql)
    export = await export_rows_to_excel(rows, base_filename="查询结果")
    print(f"[SQL-EXPORT] rows={len(rows)} file={export['filename']}")
    return {
        "fileId": export["id"],
        "filename": export["filename"],
        "downloadUrl": f"/api/files/download/{export['id']}",
        "rowCount": len(rows),
    }

# 任务：OLT 统计——按照机房统计低效 OLT 台数并导出
@app.post("/api/tasks/olt-statistics", response_model=OLTStatisticsResponse)
async def task_olt_statistics():
    """Execute OLT统计 query, export to Excel, return preview and download link."""
    sql = (
        "with OLT_and_OLT_yonghu as (\n"
        "    SELECT distinct suo_shu_qu_xian, suo_shu_ji_fang_zi_yuan_dian,\n"
        "    COUNT(*) OVER (PARTITION BY suo_shu_ji_fang_zi_yuan_dian) AS ji_fang_count,\n"
        "    SUM(CAST(onu_shu_liang as integer)) OVER (PARTITION BY suo_shu_ji_fang_zi_yuan_dian) AS yonghu_liang\n"
        "    FROM ziguan_olt_data\n"
        ")\n"
        "SELECT *, CASE WHEN ji_fang_count - CEILING(COALESCE(yonghu_liang, 0) / 4500.0) < 0 THEN 0 ELSE ji_fang_count - CEILING(COALESCE(yonghu_liang, 0) / 4500.0) END AS dixiao_OLT_taishu\n"
        "FROM OLT_and_OLT_yonghu;"
    )
    print("[TASK][OLT-STAT] executing SQL")
    rows = await execute_query_dicts(sql)
    preview = rows[:5]
    export = await export_rows_to_excel(rows, base_filename="OLT统计")
    print(f"[TASK][OLT-STAT] rows={len(rows)} file={export['filename']}")
    return {
        "preview": preview,
        "fileId": export["id"],
        "filename": export["filename"],
        "downloadUrl": f"/api/files/download/{export['id']}",
        "rowCount": len(rows),
    }

# 任务：FTTR 鉴别——支持按二级分光器或 ONU 名称查询与同机房推荐
@app.post("/api/tasks/fttr-check", response_model=FTTRCheckResponse)
async def task_fttr_check(request: FTTRCheckRequest):
    """FTTR鉴别：基于二级分光器或ONU名称。"""
    erji = (request.erjiFenGuang or "").strip()
    onu = (request.onuMingCheng or "").strip()
    if not erji and not onu:
        raise HTTPException(status_code=400, detail="请提供二级分光器名称或ONU名称")

    if onu:
        # Step 2.1: 查询ONU用户对应的二级分光器（含机房/OLT信息与在线过滤）
        sql_step1 = (
            "select C.onu_ming_cheng,\n"
            "       A.fen_guang_qi_ming_cheng as erji_fen_guang,\n"
            "       A.fen_guang_qi_ji_bie,\n"
            "       B.fen_guang_qi_ming_cheng as yiji_fen_guang,\n"
            "       B.shang_lian_she_bei as OLT_mingcheng,\n"
            "       B.shang_lian_she_bei_zhu_yong_duan_kou as OLT_PON_kou,\n"
            "       E.suo_shu_ji_fang_zi_yuan_dian as jifang,\n"
            "       A.shang_lian_she_bei_zhu_yong_duan_kou,\n"
            "       A.shang_lian_she_bei_zhu_yong_duan_kou ~ 'CG' as fenguangqi_support_open_FTTR,\n"
            "       C.zhong_duan_lei_xing,\n"
            "       C.zhong_duan_lei_xing in ('V176-20', 'HN8145XR', 'HG3142F', 'ZXHN G7611 V2', 'V175', 'V173', 'UNF130Z') as single_ONU_support_fttr\n\n"
            "from  \"wangguan_ONU_zaixianqingdan\" C\n"
            "join \"ziguan_ONU_guangmao\" D on C.onu_ming_cheng = D.xin_zeng_onu\n"
            "left join     ziguan_fenguangqi A on A.fen_guang_qi_ming_cheng = D.jie_ru_she_bei_ming_cheng\n"
            "left join \"ziguan_fenguangqi\" B\n"
            "    on B.fen_guang_qi_ming_cheng = A.shang_lian_fen_guang_qi\n"
            "left join \"ziguan_olt_data\" E on E.olt_ming_cheng = B.shang_lian_she_bei\n"
            "where A.fen_guang_qi_ji_bie = '二级分光' and C.yun_xing_zhuang_tai = '在线' and C.onu_ming_cheng = $1\n"
        )
        print(f"[TASK][FTTR][ONU] step1 onu={onu} executing SQL")
        rows_step1 = await execute_query_dicts(sql_step1, [onu])
        # If 二级分光器支持CG口，直接返回step1结果
        def _support_flag(rec: Dict[str, Any]) -> bool:
            return bool(rec.get("fenguangqi_support_open_FTTR")) or bool(rec.get("fenguangqi_support_open_fttr"))
        if rows_step1 and any(_support_flag(r) for r in rows_step1):
            rows = rows_step1
            base_name = f"FTTR鉴别_ONU_{onu}"
        else:
            # Step 2.2: 使用机房作为输入，查询同机房内满足条件的二级分光器聚合信息
            if not rows_step1 or not rows_step1[0].get("jifang"):
                # No jifang info; return step1 as fallback
                rows = rows_step1
                base_name = f"FTTR鉴别_ONU_{onu}"
            else:
                jifang = rows_step1[0]["jifang"]
                sql_step2 = (
                    "select\n"
                    "    A.fen_guang_qi_ming_cheng as erji_fen_guang,\n"
                    "    COUNT(C.onu_ming_cheng) as onu_count,\n"
                    "    A.fen_guang_qi_ji_bie,\n"
                    "    B.fen_guang_qi_ming_cheng as yiji_fen_guang,\n"
                    "    B.shang_lian_she_bei as OLT_mingcheng,\n"
                    "    B.shang_lian_she_bei_zhu_yong_duan_kou as OLT_PON_kou,\n"
                    "    E.suo_shu_ji_fang_zi_yuan_dian as jifang,\n"
                    "    A.shang_lian_she_bei_zhu_yong_duan_kou,\n"
                    "    A.shang_lian_she_bei_zhu_yong_duan_kou ~ 'CG' as fenguangqi_support_open_FTTR,\n"
                    "    STRING_AGG(C.zhong_duan_lei_xing, ', ') as zhong_duan_lei_xing_list,\n"
                    "    BOOL_OR(C.zhong_duan_lei_xing in ('V176-20', 'HN8145XR', 'HG3142F', 'ZXHN G7611 V2', 'V175', 'V173', 'UNF130Z')) as has_single_ONU_support_fttr\n\n"
                    "from \"wangguan_ONU_zaixianqingdan\" C\n"
                    "join \"ziguan_ONU_guangmao\" D on C.onu_ming_cheng = D.xin_zeng_onu\n"
                    "left join \"ziguan_fenguangqi\" A on A.fen_guang_qi_ming_cheng = D.jie_ru_she_bei_ming_cheng\n"
                    "left join \"ziguan_fenguangqi\" B\n"
                    "    on B.fen_guang_qi_ming_cheng = A.shang_lian_fen_guang_qi\n"
                    "left join \"ziguan_olt_data\" E on E.olt_ming_cheng = B.shang_lian_she_bei\n"
                    "where A.fen_guang_qi_ji_bie = '二级分光'\n"
                    "  and E.suo_shu_ji_fang_zi_yuan_dian = $1 and (A.shang_lian_she_bei_zhu_yong_duan_kou ~ 'CG') and C.zhong_duan_lei_xing not in ('V176-20', 'HN8145XR', 'HG3142F', 'ZXHN G7611 V2', 'V175', 'V173', 'UNF130Z')\n"
                    "  and C.yun_xing_zhuang_tai = '在线'\n"
                    "GROUP BY\n"
                    "    A.fen_guang_qi_ming_cheng,\n"
                    "    A.fen_guang_qi_ji_bie,\n"
                    "    B.fen_guang_qi_ming_cheng,\n"
                    "    B.shang_lian_she_bei,\n"
                    "    B.shang_lian_she_bei_zhu_yong_duan_kou,\n"
                    "    E.suo_shu_ji_fang_zi_yuan_dian,\n"
                    "    A.shang_lian_she_bei_zhu_yong_duan_kou\n"
                    "ORDER BY onu_count DESC;\n"
                )
                print(f"[TASK][FTTR][ONU] step2 jifang={jifang} executing SQL")
                rows = await execute_query_dicts(sql_step2, [jifang])
                base_name = f"FTTR鉴别_ONU_{onu}_同机房推荐"
    else:
        sql = (
            "select A.fen_guang_qi_ming_cheng as erji_fen_guang,\n"
            "       A.fen_guang_qi_ji_bie,\n"
            "       B.fen_guang_qi_ming_cheng as yiji_fen_guang,\n"
            "       B.shang_lian_she_bei as OLT_mingcheng,\n"
            "       B.shang_lian_she_bei_zhu_yong_duan_kou as OLT_PON_kou,\n"
            "       E.suo_shu_ji_fang_zi_yuan_dian as jifang,\n"
            "       A.shang_lian_she_bei_zhu_yong_duan_kou,\n"
            "       A.shang_lian_she_bei_zhu_yong_duan_kou ~ 'CG' as support_open_FTTR\n"
            "from \"ziguan_fenguangqi\" A left join \"ziguan_fenguangqi\" B on B.fen_guang_qi_ming_cheng = A.shang_lian_fen_guang_qi\n"
            "left join \"ziguan_olt_data\" E on E.olt_ming_cheng = B.shang_lian_she_bei\n"
            "where A.fen_guang_qi_ji_bie = '二级分光' and A.fen_guang_qi_ming_cheng = $1\n"
        )
        print(f"[TASK][FTTR][FGQ] erji={erji} executing SQL")
        rows = await execute_query_dicts(sql, [erji])
        base_name = f"FTTR鉴别_二级分光_{erji}"

        # 若该二级分光器不支持CG口，则基于机房进一步推荐同机房内支持的二级分光器
        def _support_fgq(rec: Dict[str, Any]) -> bool:
            return bool(rec.get("support_open_FTTR")) or bool(rec.get("support_open_fttr"))

        if (not rows) or (rows and not any(_support_fgq(r) for r in rows)):
            jifang_val = rows[0].get("jifang") if rows else None
            if jifang_val:
                sql_step2 = (
                    "select\n"
                    "    A.fen_guang_qi_ming_cheng as erji_fen_guang,\n"
                    "    COUNT(C.onu_ming_cheng) as onu_count,\n"
                    "    A.fen_guang_qi_ji_bie,\n"
                    "    B.fen_guang_qi_ming_cheng as yiji_fen_guang,\n"
                    "    B.shang_lian_she_bei as OLT_mingcheng,\n"
                    "    B.shang_lian_she_bei_zhu_yong_duan_kou as OLT_PON_kou,\n"
                    "    E.suo_shu_ji_fang_zi_yuan_dian as jifang,\n"
                    "    A.shang_lian_she_bei_zhu_yong_duan_kou,\n"
                    "    A.shang_lian_she_bei_zhu_yong_duan_kou ~ 'CG' as fenguangqi_support_open_FTTR,\n"
                    "    STRING_AGG(C.zhong_duan_lei_xing, ', ') as zhong_duan_lei_xing_list,\n"
                    "    BOOL_OR(C.zhong_duan_lei_xing in ('V176-20', 'HN8145XR', 'HG3142F', 'ZXHN G7611 V2', 'V175', 'V173', 'UNF130Z')) as has_single_ONU_support_fttr\n\n"
                    "from \"wangguan_ONU_zaixianqingdan\" C\n"
                    "join \"ziguan_ONU_guangmao\" D on C.onu_ming_cheng = D.xin_zeng_onu\n"
                    "left join \"ziguan_fenguangqi\" A on A.fen_guang_qi_ming_cheng = D.jie_ru_she_bei_ming_cheng\n"
                    "left join \"ziguan_fenguangqi\" B\n"
                    "    on B.fen_guang_qi_ming_cheng = A.shang_lian_fen_guang_qi\n"
                    "left join \"ziguan_olt_data\" E on E.olt_ming_cheng = B.shang_lian_she_bei\n"
                    "where A.fen_guang_qi_ji_bie = '二级分光'\n"
                    "  and E.suo_shu_ji_fang_zi_yuan_dian = $1 and (A.shang_lian_she_bei_zhu_yong_duan_kou ~ 'CG') and C.zhong_duan_lei_xing not in ('V176-20', 'HN8145XR', 'HG3142F', 'ZXHN G7611 V2', 'V175', 'V173', 'UNF130Z')\n"
                    "  and C.yun_xing_zhuang_tai = '在线'\n"
                    "GROUP BY\n"
                    "    A.fen_guang_qi_ming_cheng,\n"
                    "    A.fen_guang_qi_ji_bie,\n"
                    "    B.fen_guang_qi_ming_cheng,\n"
                    "    B.shang_lian_she_bei,\n"
                    "    B.shang_lian_she_bei_zhu_yong_duan_kou,\n"
                    "    E.suo_shu_ji_fang_zi_yuan_dian,\n"
                    "    A.shang_lian_she_bei_zhu_yong_duan_kou\n"
                    "ORDER BY onu_count DESC;\n"
                )
                print(f"[TASK][FTTR][FGQ] jifang-step erji={erji} jifang={jifang_val} executing SQL")
                rows = await execute_query_dicts(sql_step2, [jifang_val])
                base_name = f"FTTR鉴别_二级分光_{erji}_同机房推荐"

    preview = rows[:5]
    export = await export_rows_to_excel(rows, base_filename=base_name)
    print(f"[TASK][FTTR] rows={len(rows)} file={export['filename']}")
    return {
        "rows": rows,
        "preview": preview,
        "fileId": export["id"],
        "filename": export["filename"],
        "downloadUrl": f"/api/files/download/{export['id']}",
        "rowCount": len(rows),
    }

# 文件下载：根据登记的 file_uploads 记录安全返回本地生成文件
@app.get("/api/files/download/{file_id}")
async def download_file(file_id: str):
    """Download a generated/uploaded file by id."""
    global db_pool
    if not db_pool:
        raise HTTPException(status_code=500, detail="数据库连接不可用")
    try:
        async with db_pool.acquire() as conn:
            rec = await conn.fetchrow("select filename, path, content_type from file_uploads where id=$1", uuid.UUID(file_id))
            if not rec:
                raise HTTPException(status_code=404, detail="未找到文件")
            filename = rec["filename"]
            path = rec["path"]
            content_type = rec["content_type"] or "application/octet-stream"
            if not os.path.exists(path):
                raise HTTPException(status_code=404, detail="文件不存在或已删除")
            return FileResponse(path=path, media_type=content_type, filename=filename)
    except HTTPException:
        raise
    except Exception as e:
        raise HTTPException(status_code=500, detail=f"下载失败: {str(e)}")

# 健康检查：返回服务状态与数据库连通性
@app.get("/health")
async def health_check():
    """Health check endpoint"""
    status = {
        "status": "healthy",
        "timestamp": datetime.now().isoformat(),
        "database": "disconnected"
    }
    
    if db_pool:
        try:
            async with db_pool.acquire() as connection:
                await connection.fetchval("SELECT 1")
                status["database"] = "connected"
        except Exception:
            status["database"] = "error"
    
    return status

# File upload & listing endpoints
# 文件上传：保存到本地目录并登记数据库，按后缀自动触发后台导入
@app.post("/api/files/upload")
async def upload_file(background_tasks: BackgroundTasks, file: UploadFile = File(...)):
    global db_pool
    if not db_pool:
        raise HTTPException(status_code=500, detail="数据库连接不可用")
    try:
        storage_dir = get_storage_dir()
        upload_id = str(uuid.uuid4())
        target_path = os.path.join(storage_dir, f"{upload_id}_{file.filename}")

        size_bytes = 0
        async with aiofiles.open(target_path, 'wb') as out:
            while True:
                chunk = await file.read(1024 * 1024 * 4)  # 4MB chunks
                if not chunk:
                    break
                size_bytes += len(chunk)
                await out.write(chunk)
        # Minimal upload log
        print(f"[upload] id={upload_id} name={file.filename} type={file.content_type} size={size_bytes} path={target_path}")

        async with db_pool.acquire() as conn:
            await ensure_migrations_tables(conn)
            await conn.execute(
                "insert into file_uploads (id, filename, path, size_bytes, content_type, status) values ($1, $2, $3, $4, $5, 'uploaded')",
                uuid.UUID(upload_id), file.filename, target_path, size_bytes, file.content_type
            )

        if (file.content_type and 'csv' in (file.content_type or '').lower()) or file.filename.lower().endswith('.csv'):
            print(f"[upload] scheduling csv import id={upload_id}")
            background_tasks.add_task(import_csv_background, upload_id, target_path)
        elif file.filename.lower().endswith(('.xlsx', '.xls')):
            print(f"[upload] scheduling excel import id={upload_id}")
            background_tasks.add_task(import_excel_background, upload_id, target_path)
        else:
            print(f"[upload] no import scheduled id={upload_id} name={file.filename}")
        return {"id": upload_id, "filename": file.filename, "size": size_bytes}
    except Exception as e:
        raise HTTPException(status_code=500, detail=f"文件上传失败: {str(e)}")

# 文件列表：查询 file_uploads 表中最近的上传与导入记录
@app.get("/api/files")
async def list_files():
    global db_pool
    if not db_pool:
        raise HTTPException(status_code=500, detail="数据库连接不可用")
    try:
        async with db_pool.acquire() as conn:
            await ensure_migrations_tables(conn)
            rows = await conn.fetch(
                "select id::text as id, filename, size_bytes, content_type, status, dataset_table, rows_imported, created_at from file_uploads order by created_at desc"
            )
            return [dict(r) for r in rows]
    except Exception as e:
        raise HTTPException(status_code=500, detail=f"获取文件列表失败: {str(e)}")

# 手动触发导入：根据扩展名安排 CSV/Excel 的后台导入任务
@app.post("/api/files/import/{upload_id}")
async def trigger_import(upload_id: str, background_tasks: BackgroundTasks):
    global db_pool
    if not db_pool:
        raise HTTPException(status_code=500, detail="数据库连接不可用")
    try:
        async with db_pool.acquire() as conn:
            rec = await conn.fetchrow("select path, status from file_uploads where id=$1", uuid.UUID(upload_id))
            if not rec:
                raise HTTPException(status_code=404, detail="未找到上传记录")
            if rec["status"] in ("importing", "imported"):
                return {"message": "已在导入中或已完成"}
            file_path = rec["path"]
            print(f"[import] id={upload_id} name={file_path}")
            suffix = pathlib.Path(file_path).suffix.lower()
            if suffix == ".csv":
                print(f"[import] manual schedule csv id={upload_id}")
                background_tasks.add_task(import_csv_background, upload_id, file_path)
            elif suffix in (".xlsx", ".xls"):
                print(f"[import] manual schedule excel id={upload_id}")
                background_tasks.add_task(import_excel_background, upload_id, file_path)
            else:
                await conn.execute(
                    "update file_uploads set status='failed', updated_at=now() where id=$1",
                    uuid.UUID(upload_id),
                )
                raise HTTPException(status_code=400, detail="仅支持导入CSV或Excel文件")
            return {"message": "已触发导入"}
    except HTTPException:
        raise
    except Exception as e:
        raise HTTPException(status_code=500, detail=f"触发导入失败: {str(e)}")

if __name__ == "__main__":
    import uvicorn
    uvicorn.run(app, host="0.0.0.0", port=8000)
